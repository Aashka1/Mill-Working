from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import re
import jwt
import bcrypt
import secrets
import asyncio
import hashlib
import logging
import smtplib
from email.message import EmailMessage
from datetime import datetime, timezone, timedelta

from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ---------------- Auth utils ----------------

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=60), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="none", max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["id"] = str(user["_id"])
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# ---------------- Auth models ----------------

class RegisterBody(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "staff"

class LoginBody(BaseModel):
    email: EmailStr
    password: str

# ---------------- Auth endpoints ----------------

@api_router.post("/auth/register")
async def register(body: RegisterBody, admin: dict = Depends(require_admin)):
    """Create a user. Admin-only: this is staff onboarding, not public sign-up.

    Deliberately does not issue cookies — the calling admin stays signed in as
    themselves rather than being swapped into the account they just created.
    """
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    if len(body.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    role = body.role if body.role in ("admin", "staff") else "staff"
    doc = {"email": email, "password_hash": hash_password(body.password), "name": body.name,
           "role": role, "created_at": datetime.now(timezone.utc).isoformat()}
    res = await db.users.insert_one(doc)
    await log_audit(admin, "Created user", f"{body.name} <{email}> · {role}")
    return {"id": str(res.inserted_id), "email": email, "name": body.name, "role": role}

@api_router.post("/auth/login")
async def login(body: LoginBody, request: Request, response: Response):
    email = body.email.lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        locked_until = attempt.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Too many attempts. Try again later.")
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        count = (attempt.get("count", 0) if attempt else 0) + 1
        await db.login_attempts.update_one({"identifier": identifier},
            {"$set": {"count": count, "locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat() if count >= 5 else None}}, upsert=True)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    await db.login_attempts.delete_one({"identifier": identifier})
    uid = str(user["_id"])
    set_auth_cookies(response, create_access_token(uid, email), create_refresh_token(uid))
    return {"id": uid, "email": email, "name": user["name"], "role": user["role"]}

@api_router.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "logged out"}

@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api_router.post("/auth/refresh")
async def refresh(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        response.set_cookie("access_token", create_access_token(str(user["_id"]), user["email"]),
                            httponly=True, secure=True, samesite="none", max_age=3600, path="/")
        return {"message": "refreshed"}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ---------------- Password reset ----------------

RESET_TOKEN_TTL_MINUTES = 30
RESET_MAX_REQUESTS = 5
RESET_WINDOW_MINUTES = 60

# Deliberately identical whether or not the address exists, so this endpoint
# cannot be used to enumerate which emails have accounts.
RESET_GENERIC_REPLY = {"message": "If that email is registered, a reset link has been sent."}

class ForgotPasswordBody(BaseModel):
    email: EmailStr

class ResetPasswordBody(BaseModel):
    token: str
    password: str

def as_utc(value: Optional[datetime]) -> Optional[datetime]:
    """Mongo hands back naive datetimes; comparing those to an aware `now`
    raises TypeError. Normalise before any comparison."""
    if value is None:
        return None
    return value if value.tzinfo else value.replace(tzinfo=timezone.utc)

def hash_reset_token(token: str) -> str:
    """Store only a digest, so a leaked database dump yields no usable tokens.

    Plain SHA-256 rather than bcrypt: the token is 32 bytes of CSPRNG output,
    so there is no low-entropy secret for an attacker to grind against.
    """
    return hashlib.sha256(token.encode("utf-8")).hexdigest()

def smtp_settings() -> Optional[dict]:
    """SMTP config, or None when it is not fully configured."""
    host = os.environ.get("SMTP_HOST")
    user = os.environ.get("SMTP_USER")
    password = os.environ.get("SMTP_PASS")
    if not (host and user and password):
        return None
    return {"host": host, "port": int(os.environ.get("SMTP_PORT", "587")), "user": user,
            "password": password, "sender": os.environ.get("SMTP_FROM", user)}

def send_email_blocking(cfg: dict, to: str, subject: str, body: str):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = cfg["sender"]
    msg["To"] = to
    msg.set_content(body)
    if cfg["port"] == 465:
        with smtplib.SMTP_SSL(cfg["host"], cfg["port"], timeout=20) as smtp:
            smtp.login(cfg["user"], cfg["password"])
            smtp.send_message(msg)
    else:
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=20) as smtp:
            smtp.starttls()
            smtp.login(cfg["user"], cfg["password"])
            smtp.send_message(msg)

def public_base_url(request: Request) -> str:
    """Origin to build reset links from, honouring the proxy Render sits behind."""
    configured = os.environ.get("PUBLIC_URL") or os.environ.get("FRONTEND_URL")
    if configured:
        return configured.rstrip("/")
    proto = request.headers.get("x-forwarded-proto", request.url.scheme)
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc
    return f"{proto}://{host}"

@api_router.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordBody, request: Request):
    email = body.email.lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    now = datetime.now(timezone.utc)

    # Throttle per IP+email. Kept separate from login_attempts so that asking
    # for a reset can never lock someone out of signing in normally.
    attempt = await db.reset_attempts.find_one({"identifier": identifier})
    if attempt:
        window_start = as_utc(attempt.get("window_start"))
        if window_start and window_start > now - timedelta(minutes=RESET_WINDOW_MINUTES):
            if attempt.get("count", 0) >= RESET_MAX_REQUESTS:
                raise HTTPException(status_code=429, detail="Too many reset requests. Try again later.")
            await db.reset_attempts.update_one({"identifier": identifier}, {"$inc": {"count": 1}})
        else:
            await db.reset_attempts.update_one({"identifier": identifier},
                {"$set": {"count": 1, "window_start": now}})
    else:
        await db.reset_attempts.insert_one({"identifier": identifier, "count": 1, "window_start": now})

    user = await db.users.find_one({"email": email})
    if not user:
        return RESET_GENERIC_REPLY

    # One live token per user: issuing a new link retires any earlier one.
    await db.password_resets.delete_many({"user_id": str(user["_id"])})
    token = secrets.token_urlsafe(32)
    await db.password_resets.insert_one({
        "token_hash": hash_reset_token(token),
        "user_id": str(user["_id"]),
        "email": email,
        "expires_at": now + timedelta(minutes=RESET_TOKEN_TTL_MINUTES),
        "created_at": now,
    })

    link = f"{public_base_url(request)}/reset-password?token={token}"
    cfg = smtp_settings()
    if cfg:
        try:
            await asyncio.to_thread(
                send_email_blocking, cfg, email, "Reset your AgriMill password",
                f"Hello {user.get('name', '')},\n\n"
                f"Use the link below to set a new AgriMill password. "
                f"It expires in {RESET_TOKEN_TTL_MINUTES} minutes and can only be used once.\n\n"
                f"{link}\n\n"
                f"If you did not request this, ignore this email — your password stays unchanged.\n")
        except Exception:
            # Never surface delivery failures: doing so would confirm the
            # address exists. Logged for the operator instead.
            logger.exception("Failed to send password reset email to %s", email)
    else:
        # No SMTP configured: the operator reads the link from the service logs
        # and passes it on. See DEPLOY.md.
        logger.warning("SMTP not configured. Password reset link for %s: %s", email, link)

    return RESET_GENERIC_REPLY

@api_router.post("/auth/reset-password")
async def reset_password(body: ResetPasswordBody):
    if len(body.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    record = await db.password_resets.find_one({"token_hash": hash_reset_token(body.token)})
    if not record:
        raise HTTPException(status_code=400, detail="This reset link is invalid or has already been used")
    expires_at = as_utc(record.get("expires_at"))
    if expires_at and expires_at < datetime.now(timezone.utc):
        await db.password_resets.delete_one({"_id": record["_id"]})
        raise HTTPException(status_code=400, detail="This reset link has expired. Request a new one.")
    try:
        oid = ObjectId(record["user_id"])
    except Exception:
        raise HTTPException(status_code=400, detail="This reset link is invalid or has already been used")
    user = await db.users.find_one({"_id": oid})
    if not user:
        await db.password_resets.delete_one({"_id": record["_id"]})
        raise HTTPException(status_code=400, detail="This reset link is invalid or has already been used")

    await db.users.update_one({"_id": oid}, {"$set": {
        "password_hash": hash_password(body.password),
        # Stops the startup admin re-seed from reverting this password on the
        # next boot. See the startup() comment.
        "password_self_managed": True,
    }})
    # Single use, and clear any lockout the forgotten password caused.
    await db.password_resets.delete_many({"user_id": record["user_id"]})
    await db.login_attempts.delete_many({"identifier": {"$regex": f":{re.escape(user['email'])}$"}})
    await log_audit(None, "Password reset", user["email"])
    return {"message": "Password updated. You can sign in now."}

@api_router.get("/users")
async def list_users(user: dict = Depends(require_admin)):
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    for u in users:
        u["id"] = str(u["_id"]); u.pop("_id", None)
    return users

@api_router.delete("/users/{uid}")
async def delete_user(uid: str, admin: dict = Depends(require_admin)):
    try:
        oid = ObjectId(uid)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user id")
    if uid == admin["id"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    target = await db.users.find_one({"_id": oid})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    # Never leave the mill without an admin who can add users back.
    if target.get("role") == "admin" and await db.users.count_documents({"role": "admin"}) <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last admin")
    await db.users.delete_one({"_id": oid})
    await log_audit(admin, "Deleted user", f'{target.get("name")} <{target.get("email")}>')
    return {"message": "deleted"}

# ---------------- Generic helpers ----------------

def clean(doc: dict) -> dict:
    doc = dict(doc)
    doc.pop("_id", None)
    return doc

def now_iso():
    return datetime.now(timezone.utc).isoformat()

async def next_invoice_number():
    count = await db.invoices.count_documents({})
    return f"INV-{datetime.now().year}-{count + 1:04d}"

# ---------------- Products / Inventory ----------------

class ProductBody(BaseModel):
    name: str
    category: str
    unit: str = "kg"
    current_stock: float = 0
    rate: float = 0
    cost_per_unit: float = 0
    low_stock_threshold: float = 50

def product_key(name: str) -> str:
    """Normalised product name. 'Atta', 'atta' and ' Atta ' are one product.

    Stock is moved by name in several places (grinding fees, exchange, oil
    retention), so two rows sharing a name would make those updates land on
    whichever row Mongo happened to return first. Names are kept unique.
    """
    return " ".join((name or "").split()).lower()

@api_router.get("/products")
async def get_products(user: dict = Depends(get_current_user)):
    return [clean(p) for p in await db.products.find().sort("name", 1).to_list(1000)]

@api_router.post("/products")
async def create_product(body: ProductBody, user: dict = Depends(get_current_user)):
    key = product_key(body.name)
    if not key:
        raise HTTPException(status_code=400, detail="Enter a product name")

    existing = await db.products.find_one({"name_key": key})
    if existing:
        # Same name in a different category is ambiguous rather than a top-up:
        # refuse instead of silently filing stock under the wrong category.
        if existing.get("category") != body.category:
            raise HTTPException(status_code=400, detail=(
                f'"{existing["name"]}" already exists under category '
                f'"{existing.get("category")}". Pick that category to add stock to it, '
                f"or give this product a different name."))
        if existing.get("unit") != body.unit:
            raise HTTPException(status_code=400, detail=(
                f'"{existing["name"]}" is measured in {existing.get("unit")}, not {body.unit}. '
                f"Edit the product if the unit needs to change."))
        # Adding an existing product tops up its stock rather than creating a
        # second row, and carries the purchase cost into the running average.
        added = body.current_stock
        if added:
            await add_stock_with_cost(existing["id"], added, added * (body.rate or existing.get("rate", 0)))
        updates = {}
        if body.rate:
            updates["rate"] = body.rate
        if body.low_stock_threshold:
            updates["low_stock_threshold"] = body.low_stock_threshold
        if updates:
            await db.products.update_one({"id": existing["id"]}, {"$set": updates})
        merged = await db.products.find_one({"id": existing["id"]})
        await log_audit(user, "Added stock", f'{merged["name"]}: +{added} {merged.get("unit")} (now {merged.get("current_stock")})')
        return {**clean(merged), "merged": True}

    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "name_key": key, "created_at": now_iso()}
    await db.products.insert_one(doc)
    await log_audit(user, "Created product", f'{body.name} ({body.category})')
    return {**clean(doc), "merged": False}

@api_router.put("/products/{pid}")
async def update_product(pid: str, body: ProductBody, user: dict = Depends(get_current_user)):
    old = await db.products.find_one({"id": pid})
    if not old:
        raise HTTPException(status_code=404, detail="Product not found")
    key = product_key(body.name)
    if not key:
        raise HTTPException(status_code=400, detail="Enter a product name")
    # Renaming onto another product would recreate the duplicate this guards against.
    clash = await db.products.find_one({"name_key": key, "id": {"$ne": pid}})
    if clash:
        raise HTTPException(status_code=400, detail=f'Another product is already named "{clash["name"]}"')
    # Never write stock or cost from this endpoint. Both are running totals that
    # sales, purchases, grinding, exchange and production move with $inc; an
    # absolute $set here would erase any of their updates that landed while the
    # edit form was open. cost_per_unit is the worse of the two — the form does
    # not send it, so the model default of 0 would wipe the cost basis that
    # profit reporting depends on. Stock corrections go through /adjust.
    fields = body.model_dump(exclude={"current_stock", "cost_per_unit"})
    await db.products.update_one({"id": pid}, {"$set": {**fields, "name_key": key}})
    await log_audit(user, "Edited product", f'{body.name} ({body.category}, {body.unit}, Rs {body.rate}/{body.unit})')
    return clean(await db.products.find_one({"id": pid}))

class StockAdjustBody(BaseModel):
    delta: float
    reason: str = ""

@api_router.post("/products/{pid}/adjust")
async def adjust_product_stock(pid: str, body: StockAdjustBody, user: dict = Depends(get_current_user)):
    """Relative stock correction — spillage, recount, damage.

    Relative rather than absolute so a concurrent sale or purchase is added to
    rather than overwritten, and always audited so a manual correction is
    distinguishable from a sale in the trail.
    """
    p = await db.products.find_one({"id": pid})
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    if not body.delta:
        raise HTTPException(status_code=400, detail="Enter a non-zero amount")
    new_stock = round(p.get("current_stock", 0) + body.delta, 3)
    if new_stock < 0:
        raise HTTPException(status_code=400, detail=(
            f'Only {p.get("current_stock", 0)} {p.get("unit")} in stock — cannot remove {abs(body.delta)}'))
    await db.products.update_one({"id": pid}, {"$inc": {"current_stock": body.delta}})
    # Recorded as a dated movement, not only in the audit text. The item report
    # reconstructs opening stock by walking back from today's figure, so an
    # adjustment with no date would silently unbalance every period before it.
    await db.stock_adjustments.insert_one({
        "id": str(uuid.uuid4()), "product_id": pid, "product_name": p["name"],
        "delta": round(body.delta, 3), "reason": body.reason or "",
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"), "created_at": now_iso()})
    await log_audit(user, "Adjusted stock", (
        f'{p["name"]}: {p.get("current_stock", 0)} → {new_stock} {p.get("unit")} '
        f'({body.delta:+g}){" · " + body.reason if body.reason else ""}'))
    return clean(await db.products.find_one({"id": pid}))

@api_router.delete("/products/{pid}")
async def delete_product(pid: str, user: dict = Depends(require_admin)):
    p = await db.products.find_one({"id": pid})
    await db.products.delete_one({"id": pid})
    await log_audit(user, "Deleted product", (p or {}).get("name", pid))
    return {"message": "deleted"}

# ---------------- Purchases ----------------

class PurchaseBody(BaseModel):
    date: str
    supplier_id: Optional[str] = None
    supplier_name: str
    # Blank when buying something not yet in the catalogue; the product is then
    # resolved (or created) from product_name.
    product_id: str = ""
    product_name: str
    quantity: float
    rate: float
    payment_status: str = "Paid"
    # Blank means "settled in full"; any number records a part payment.
    amount_paid: Optional[float] = None
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None
    # Only used when a purchase creates the product.
    unit: Optional[str] = None
    category: Optional[str] = None


async def resolve_purchase_product(body: "PurchaseBody") -> dict:
    """The product a purchase stocks, creating it when the item is new.

    Buying something always has to land somewhere. Matching on id alone meant a
    stale or merged-away id silently moved no stock at all — the purchase was
    recorded and the shelf figure never changed. So fall back to the name, and
    if the mill has genuinely bought something new, add it to the catalogue with
    this quantity rather than dropping it.
    """
    if body.product_id:
        p = await db.products.find_one({"id": body.product_id})
        if p:
            return p
    name = (body.product_name or "").strip()
    key = product_key(name)
    if not key:
        raise HTTPException(status_code=400, detail="Select a product or enter a name for it")
    p = await db.products.find_one({"name_key": key})
    if p:
        return p
    doc = {"id": str(uuid.uuid4()), "name": name, "name_key": key,
           "category": (body.category or "Other").strip() or "Other",
           "unit": (body.unit or "kg").strip() or "kg",
           "current_stock": 0, "cost_per_unit": 0, "rate": 0,
           "low_stock_threshold": 0, "created_at": now_iso()}
    await db.products.insert_one(doc)
    logger.info("Purchase created new product %r", name)
    return doc

@api_router.get("/purchases")
async def get_purchases(user: dict = Depends(get_current_user)):
    return [clean(p) for p in await db.purchases.find().sort("date", -1).to_list(2000)]

@api_router.post("/purchases")
async def create_purchase(body: PurchaseBody, user: dict = Depends(get_current_user)):
    total = round(body.quantity * body.rate, 2)
    prod = await resolve_purchase_product(body)
    fields = body.model_dump(exclude={"unit", "category"})
    # Store the resolved ids so the row always points at the product it stocked.
    fields.update({"product_id": prod["id"], "product_name": prod["name"]})
    doc = {"id": str(uuid.uuid4()), **fields, "total": total, "created_at": now_iso()}
    await db.purchases.insert_one(doc)
    await add_stock_with_cost(prod["id"], body.quantity, total)
    received = total if body.amount_paid is None and body.payment_status == "Paid" else (body.amount_paid or 0)
    await add_credit("supplier", body.supplier_name, min(received, total), body.date, doc["id"], f"Purchase {body.product_name}", mode=body.payment_mode, bank_id=body.bank_id)
    await sync_payment_state("purchases", doc["id"])
    doc = await db.purchases.find_one({"id": doc["id"]})
    await log_audit(user, "Created purchase", f'{body.supplier_name} · {prod["name"]} {body.quantity} {prod.get("unit", "kg")} · Rs {total}')
    return clean(doc)

@api_router.put("/purchases/{pid}")
async def edit_purchase(pid: str, body: PurchaseBody, user: dict = Depends(get_current_user)):
    old = await db.purchases.find_one({"id": pid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    # Reverse the original stock movement before applying the corrected one, so
    # editing a purchase cannot double-count or strand stock on the old product.
    await add_stock_with_cost(old["product_id"], -old["quantity"], -old.get("total", 0))
    total = round(body.quantity * body.rate, 2)
    prod = await resolve_purchase_product(body)
    fields = body.model_dump(exclude={"unit", "category"})
    fields.update({"product_id": prod["id"], "product_name": prod["name"]})
    await db.purchases.update_one({"id": pid}, {"$set": {**fields, "total": total}})
    await add_stock_with_cost(prod["id"], body.quantity, total)
    await retag_bill_payments(pid, body.payment_mode, body.bank_id)
    await sync_payment_state("purchases", pid)
    await log_audit(user, "Edited purchase", f'{body.supplier_name} · {prod["name"]} {body.quantity} {prod.get("unit", "kg")} · Rs {total}')
    return clean(await db.purchases.find_one({"id": pid}))

@api_router.delete("/purchases/{pid}")
async def delete_purchase(pid: str, user: dict = Depends(require_admin)):
    p = await db.purchases.find_one({"id": pid})
    if p:
        await db.products.update_one({"id": p["product_id"]}, {"$inc": {"current_stock": -p["quantity"]}})
        await db.purchases.delete_one({"id": pid})
        for row in await db.payments.find({"ref_id": pid}).to_list(500):
            await unpost_bank_txn(row["id"])
        await db.payments.delete_many({"ref_id": pid})
    return {"message": "deleted"}

# ---------------- Sales ----------------

class SaleBody(BaseModel):
    date: str
    customer_id: Optional[str] = None
    customer_name: str
    product_id: str
    product_name: str
    quantity: float
    price: float
    payment_status: str = "Paid"
    amount_paid: Optional[float] = None
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None

async def product_cost(product_id: str) -> float:
    p = await db.products.find_one({"id": product_id})
    return round((p or {}).get("cost_per_unit", 0) or 0, 4)

def sale_cogs(sale: dict, cost_by_name: dict) -> float:
    """Cost of the goods in one sale.

    Uses the cost captured on the sale when present. Sales written before that
    snapshot existed fall back to the product's current average cost, which is
    what the old reports did and is the best available for historical rows.
    """
    if sale.get("cogs") is not None:
        return sale.get("cogs", 0) or 0
    unit = sale.get("unit_cost")
    if unit is None:
        unit = cost_by_name.get(sale.get("product_name"), 0)
    return round((sale.get("quantity", 0) or 0) * (unit or 0), 2)

@api_router.get("/sales")
async def get_sales(user: dict = Depends(get_current_user)):
    return [clean(s) for s in await db.sales.find().sort("date", -1).to_list(2000)]

@api_router.post("/sales")
async def create_sale(body: SaleBody, user: dict = Depends(get_current_user)):
    total = round(body.quantity * body.price, 2)
    inv = await next_invoice_number()
    # Snapshot what this stock cost at the moment it left the shop. cost_per_unit
    # on the product is a running weighted average, so reading it later would let
    # a future purchase silently rewrite the profit on a sale already made.
    unit_cost = await product_cost(body.product_id)
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "total": total, "invoice_number": inv,
           "unit_cost": unit_cost, "cogs": round(body.quantity * unit_cost, 2), "created_at": now_iso()}
    await db.sales.insert_one(doc)
    await db.products.update_one({"id": body.product_id}, {"$inc": {"current_stock": -body.quantity}})
    await db.invoices.insert_one({"id": str(uuid.uuid4()), "invoice_number": inv, "type": "Sale",
        "ref_id": doc["id"], "customer_name": body.customer_name, "date": body.date,
        "total": total, "payment_status": body.payment_status, "created_at": now_iso()})
    received = total if body.amount_paid is None and body.payment_status == "Paid" else (body.amount_paid or 0)
    await add_credit("customer", body.customer_name, min(received, total), body.date, doc["id"], f"Sale {inv}", mode=body.payment_mode, bank_id=body.bank_id)
    await sync_payment_state("sales", doc["id"])
    doc = await db.sales.find_one({"id": doc["id"]})
    await log_audit(user, "Created sale", f"{body.customer_name} · {body.product_name} {body.quantity} {await product_unit(body.product_id)} · Rs {total}")
    return clean(doc)

@api_router.delete("/sales/{sid}")
async def delete_sale(sid: str, user: dict = Depends(require_admin)):
    s = await db.sales.find_one({"id": sid})
    if s:
        await db.products.update_one({"id": s["product_id"]}, {"$inc": {"current_stock": s["quantity"]}})
        await db.sales.delete_one({"id": sid})
        await db.invoices.delete_one({"ref_id": sid})
        for row in await db.payments.find({"ref_id": sid}).to_list(500):
            await unpost_bank_txn(row["id"])
        await db.payments.delete_many({"ref_id": sid})
    return {"message": "deleted"}

# ---------------- Grinding ----------------

class GrindingBody(BaseModel):
    date: str
    customer_id: Optional[str] = None
    customer_name: str
    # wheat_weight predates the mill grinding anything but wheat; it now means
    # "weight of whatever was brought in", kept under the old name so existing
    # records and reports stay readable.
    grain_type: str = "Wheat"
    output_product: str = "Atta"
    wheat_weight: float
    washed: bool = True
    loss_percent: float = 2.5
    charge_per_kg: float = 0
    # Cash / UPI / Bank / NEFT / RTGS / IMPS / Cheque, or paid in kind with
    # Flour Deduction or Grain Deduction.
    payment_method: str = "Cash"
    grain_fee_kg: float = 0
    # Flour deduction: how much to keep back, and on what basis.
    deduction_basis: str = "Percent"
    deduction_percent: Optional[float] = None
    deduction_weight: Optional[float] = None
    # Grain deduction: what came in instead of money.
    grain_item: str = ""
    grain_qty: Optional[float] = None
    grain_value: Optional[float] = None
    payment_status: str = "Pending"
    amount_paid: Optional[float] = None
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None

@api_router.get("/grinding")
async def get_grinding(user: dict = Depends(get_current_user)):
    return [clean(g) for g in await db.grinding.find().sort("date", -1).to_list(2000)]

@api_router.post("/grinding")
async def create_grinding(body: GrindingBody, user: dict = Depends(get_current_user)):
    doc = await build_grinding_doc(body)
    await db.grinding.insert_one(doc)
    await apply_grinding_effects(doc, 1)
    await db.invoices.insert_one({"id": str(uuid.uuid4()), "invoice_number": doc["invoice_number"], "type": "Grinding",
        "ref_id": doc["id"], "customer_name": doc["customer_name"], "date": doc["date"],
        "total": doc["total_charge"], "payment_status": doc["payment_status"], "created_at": now_iso()})
    charge = doc["total_charge"]
    method = normalise_method(doc.get("payment_method"))
    if settled_in_kind(method):
        # The flour or grain kept back is the payment, so the bill is settled in
        # full. Crediting only what the goods valued at would leave small
        # phantom dues whenever the rate and the charge did not divide evenly.
        kind_mode = "Flour" if method == FLOUR_DEDUCTION else "Grain"
        detail = (f'{doc.get("deducted_flour", 0)} kg {grinding_output_name(doc)} kept'
                  if method == FLOUR_DEDUCTION
                  else f'{doc.get("grain_qty", 0)} kg {doc.get("grain_item") or "grain"} received')
        await add_credit("customer", doc["customer_name"], charge, doc["date"], doc["id"],
                         f'Grinding {doc["invoice_number"]} · {detail}', mode=kind_mode)
    else:
        received = charge if body.amount_paid is None and doc["payment_status"] == "Paid" else (body.amount_paid or 0)
        await add_credit("customer", doc["customer_name"], min(received, charge), doc["date"], doc["id"], f"Grinding {doc['invoice_number']}", mode=body.payment_mode or method, bank_id=body.bank_id)
    await sync_payment_state("grinding", doc["id"])
    doc = await db.grinding.find_one({"id": doc["id"]})
    return clean(doc)

@api_router.delete("/grinding/{gid}")
async def delete_grinding(gid: str, user: dict = Depends(require_admin)):
    g = await db.grinding.find_one({"id": gid})
    if g:
        await apply_grinding_effects(g, -1)
        await unpost_material_ledger(gid)
        await db.grinding.delete_one({"id": gid})
        await db.invoices.delete_one({"ref_id": gid})
        for row in await db.payments.find({"ref_id": gid}).to_list(500):
            await unpost_bank_txn(row["id"])
        await db.payments.delete_many({"ref_id": gid})
    return {"message": "deleted"}

# ---------------- Oil Extraction ----------------

class OilBody(BaseModel):
    date: str
    customer_id: Optional[str] = None
    customer_name: str
    seed_type: str
    quantity_received: float
    oil_extracted: float
    oil_cake_produced: float = 0
    charge: float = 0
    payment_method: str = "Cash"
    retained_oil: float = 0
    retained_cake: float = 0
    # Cake the customer sells to the shop, priced per kg. Distinct from
    # retained_cake: that is cake kept as the processing fee and arrives free,
    # whereas this is a purchase, so it carries a cost and its value comes off
    # the customer's bill.
    cake_sold_to_shop: float = 0
    cake_rate: float = 0
    payment_status: str = "Pending"
    amount_paid: Optional[float] = None
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None

@api_router.get("/oil")
async def get_oil(user: dict = Depends(get_current_user)):
    return [clean(o) for o in await db.oil.find().sort("date", -1).to_list(2000)]

@api_router.post("/oil")
async def create_oil(body: OilBody, user: dict = Depends(get_current_user)):
    doc = await build_oil_doc(body)
    await db.oil.insert_one(doc)
    await apply_oil_effects(doc, 1)
    await db.invoices.insert_one({"id": str(uuid.uuid4()), "invoice_number": doc["invoice_number"], "type": "Oil Extraction",
        "ref_id": doc["id"], "customer_name": doc["customer_name"], "date": doc["date"],
        "total": doc["total"], "payment_status": doc["payment_status"], "created_at": now_iso()})
    charge = doc["total"]
    if charge < 0:
        # The cake is worth more than the grinding: hand the difference over.
        # Recorded at once, because the customer leaves with the cash.
        await add_credit("customer", doc["customer_name"], charge, doc["date"], doc["id"],
                         f"Paid to customer · Oil {doc['invoice_number']}", kind="refund", mode=body.payment_mode, bank_id=body.bank_id)
    else:
        received = charge if body.amount_paid is None and doc["payment_status"] == "Paid" else (body.amount_paid or 0)
        await add_credit("customer", doc["customer_name"], min(received, charge), doc["date"], doc["id"], f"Oil {doc['invoice_number']}", mode=body.payment_mode, bank_id=body.bank_id)
    await sync_payment_state("oil", doc["id"])
    doc = await db.oil.find_one({"id": doc["id"]})
    return clean(doc)

@api_router.delete("/oil/{oid}")
async def delete_oil(oid: str, user: dict = Depends(require_admin)):
    o = await db.oil.find_one({"id": oid})
    if o:
        await apply_oil_effects(o, -1)
        await db.oil.delete_one({"id": oid})
        await db.invoices.delete_one({"ref_id": oid})
        for row in await db.payments.find({"ref_id": oid}).to_list(500):
            await unpost_bank_txn(row["id"])
        await db.payments.delete_many({"ref_id": oid})
    return {"message": "deleted"}

# ---------------- Expenses ----------------

class ExpenseBody(BaseModel):
    date: str
    category: str
    description: str = ""
    amount: float

@api_router.get("/expenses")
async def get_expenses(user: dict = Depends(get_current_user)):
    return [clean(e) for e in await db.expenses.find().sort("date", -1).to_list(2000)]

@api_router.post("/expenses")
async def create_expense(body: ExpenseBody, user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_iso()}
    await db.expenses.insert_one(doc)
    return clean(doc)

@api_router.delete("/expenses/{eid}")
async def delete_expense(eid: str, user: dict = Depends(require_admin)):
    await db.expenses.delete_one({"id": eid})
    return {"message": "deleted"}

# ---------------- Customers ----------------

class CustomerBody(BaseModel):
    name: str
    phone: str = ""
    address: str = ""
    # Optional on purpose: a walk-in customer has a name and nothing else, and
    # demanding more would push staff into inventing values.
    gstin: str = ""
    pan_aadhaar: str = ""
    # What the party already owed when they were first entered, so a balance
    # carried over from the old books is not lost.
    opening_balance: float = 0
    credit_limit: float = 0

@api_router.get("/customers")
async def get_customers(user: dict = Depends(get_current_user)):
    customers = [clean(c) for c in await db.customers.find().sort("name", 1).to_list(2000)]
    for c in customers:
        debit = 0.0
        for coll, field in [("sales", "total"), ("grinding", "total_charge"), ("oil", "total"),
                            ("exchanges", "grinding_charge")]:
            docs = await db[coll].find({"customer_name": c["name"]}).to_list(2000)
            debit += sum(d.get(field, 0) for d in docs)
        credit = sum(p.get("amount", 0) for p in await db.payments.find({"party_type": "customer", "party_name": c["name"]}).to_list(2000))
        c["outstanding"] = round(debit + (c.get("opening_balance", 0) or 0) - credit, 2)
    return customers

@api_router.post("/customers")
async def create_customer(body: CustomerBody, user: dict = Depends(get_current_user)):
    """Create a customer, or return the existing one under that name.

    Called inline from the transaction screens, where the operator is halfway
    through a bill. Returning the match rather than erroring means a name typed
    twice never splits into two ledgers, and never blocks the bill in progress.
    """
    name = " ".join((body.name or "").split())
    if not name:
        raise HTTPException(status_code=400, detail="Enter a name")
    key = party_key(name)
    existing = await db.customers.find_one({"party_key": key})
    if existing:
        out = clean(existing)
        out["existing"] = True
        return out
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "name": name,
           "party_key": key, "created_at": now_iso()}
    await db.customers.insert_one(doc)
    await log_audit(user, "Added customer", name)
    out = clean(doc)
    out["existing"] = False
    return out

@api_router.put("/customers/{cid}")
async def update_customer(cid: str, body: CustomerBody, user: dict = Depends(get_current_user)):
    old = await db.customers.find_one({"id": cid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    name = " ".join((body.name or "").split())
    if not name:
        raise HTTPException(status_code=400, detail="Enter a name")
    key = party_key(name)
    clash = await db.customers.find_one({"party_key": key, "id": {"$ne": cid}})
    if clash:
        raise HTTPException(status_code=400, detail=f'Another customer is already called "{clash["name"]}"')
    await db.customers.update_one({"id": cid}, {"$set": {**body.model_dump(), "name": name, "party_key": key}})
    # Every transaction stores the party by name, so a rename has to carry them
    # with it or the ledger silently empties.
    if old.get("name") and old["name"] != name:
        for c, field in (('sales', 'customer_name'), ('grinding', 'customer_name'), ('oil', 'customer_name'), ('exchanges', 'customer_name')):
            await db[c].update_many({field: old["name"]}, {"$set": {field: name}})
        await db.payments.update_many({"party_type": "customer", "party_name": old["name"]},
                                      {"$set": {"party_name": name}})
        await log_audit(user, "Renamed customer", f'{old["name"]} to {name}')
    return clean(await db.customers.find_one({"id": cid}))

@api_router.get("/customers/{cid}/history")
async def customer_history(cid: str, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one({"id": cid})
    if not c:
        raise HTTPException(status_code=404, detail="Not found")
    name = c["name"]
    return {
        "sales": [clean(x) for x in await db.sales.find({"customer_name": name}).to_list(1000)],
        "grinding": [clean(x) for x in await db.grinding.find({"customer_name": name}).to_list(1000)],
        "oil": [clean(x) for x in await db.oil.find({"customer_name": name}).to_list(1000)],
    }

@api_router.delete("/customers/{cid}")
async def delete_customer(cid: str, user: dict = Depends(require_admin)):
    party = await db.customers.find_one({"id": cid})
    if not party:
        raise HTTPException(status_code=404, detail="Not found")
    # Transactions reference the party by name, so removing one with history
    # would leave those records pointing at nobody.
    for c, field in (("sales", "customer_name"), ("grinding", "customer_name"), ("oil", "customer_name"), ("exchanges", "customer_name")):
        if await db[c].count_documents({field: party["name"]}):
            raise HTTPException(status_code=400,
                detail=f'{party["name"]} has transactions and cannot be deleted. Edit the record instead.')
    await db.customers.delete_one({"id": cid})
    await log_audit(user, "Deleted customer", party["name"])
    return {"message": "deleted"}

# ---------------- Suppliers ----------------

class SupplierBody(BaseModel):
    name: str
    phone: str = ""
    address: str = ""
    # Optional on purpose: a walk-in customer has a name and nothing else, and
    # demanding more would push staff into inventing values.
    gstin: str = ""
    pan_aadhaar: str = ""
    # What the party already owed when they were first entered, so a balance
    # carried over from the old books is not lost.
    opening_balance: float = 0
    credit_limit: float = 0

@api_router.get("/suppliers")
async def get_suppliers(user: dict = Depends(get_current_user)):
    suppliers = [clean(s) for s in await db.suppliers.find().sort("name", 1).to_list(2000)]
    for s in suppliers:
        docs = await db.purchases.find({"supplier_name": s["name"]}).to_list(2000)
        debit = sum(d.get("total", 0) for d in docs)
        credit = sum(p.get("amount", 0) for p in await db.payments.find({"party_type": "supplier", "party_name": s["name"]}).to_list(2000))
        s["outstanding"] = round(debit + (s.get("opening_balance", 0) or 0) - credit, 2)
    return suppliers

@api_router.post("/suppliers")
async def create_supplier(body: SupplierBody, user: dict = Depends(get_current_user)):
    """Create a supplier, or return the existing one under that name.

    Called inline from the transaction screens, where the operator is halfway
    through a bill. Returning the match rather than erroring means a name typed
    twice never splits into two ledgers, and never blocks the bill in progress.
    """
    name = " ".join((body.name or "").split())
    if not name:
        raise HTTPException(status_code=400, detail="Enter a name")
    key = party_key(name)
    existing = await db.suppliers.find_one({"party_key": key})
    if existing:
        out = clean(existing)
        out["existing"] = True
        return out
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "name": name,
           "party_key": key, "created_at": now_iso()}
    await db.suppliers.insert_one(doc)
    await log_audit(user, "Added supplier", name)
    out = clean(doc)
    out["existing"] = False
    return out

@api_router.put("/suppliers/{sid}")
async def update_supplier(sid: str, body: SupplierBody, user: dict = Depends(get_current_user)):
    old = await db.suppliers.find_one({"id": sid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    name = " ".join((body.name or "").split())
    if not name:
        raise HTTPException(status_code=400, detail="Enter a name")
    key = party_key(name)
    clash = await db.suppliers.find_one({"party_key": key, "id": {"$ne": sid}})
    if clash:
        raise HTTPException(status_code=400, detail=f'Another supplier is already called "{clash["name"]}"')
    await db.suppliers.update_one({"id": sid}, {"$set": {**body.model_dump(), "name": name, "party_key": key}})
    # Every transaction stores the party by name, so a rename has to carry them
    # with it or the ledger silently empties.
    if old.get("name") and old["name"] != name:
        for c, field in (('purchases', 'supplier_name'),):
            await db[c].update_many({field: old["name"]}, {"$set": {field: name}})
        await db.payments.update_many({"party_type": "supplier", "party_name": old["name"]},
                                      {"$set": {"party_name": name}})
        await log_audit(user, "Renamed supplier", f'{old["name"]} to {name}')
    return clean(await db.suppliers.find_one({"id": sid}))

@api_router.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, user: dict = Depends(require_admin)):
    party = await db.suppliers.find_one({"id": sid})
    if not party:
        raise HTTPException(status_code=404, detail="Not found")
    # Transactions reference the party by name, so removing one with history
    # would leave those records pointing at nobody.
    for c, field in (("purchases", "supplier_name"),):
        if await db[c].count_documents({field: party["name"]}):
            raise HTTPException(status_code=400,
                detail=f'{party["name"]} has transactions and cannot be deleted. Edit the record instead.')
    await db.suppliers.delete_one({"id": sid})
    await log_audit(user, "Deleted supplier", party["name"])
    return {"message": "deleted"}

# ---------------- Invoices ----------------

@api_router.get("/invoices")
async def get_invoices(user: dict = Depends(get_current_user)):
    return [clean(i) for i in await db.invoices.find().sort("created_at", -1).to_list(2000)]

async def product_unit(product_id: str, default: str = "kg") -> str:
    """Unit a product is measured in. Sales do not store it, so read the catalogue.

    Matters on customer-facing invoices: packing is sold by the bag and oil by
    the litre, and printing either as "kg" is simply wrong.
    """
    if not product_id:
        return default
    p = await db.products.find_one({"id": product_id})
    return (p or {}).get("unit") or default

def grinding_invoice_items(g: dict) -> list:
    """Invoice lines for a grinding job.

    The charge is always its own line, whatever settled it. When flour was kept
    back, the deduction and what the customer actually walks out with are shown
    underneath — otherwise the bill says one weight and the customer carries a
    different one home.
    """
    out_name = grinding_output_name(g)
    items = [{"desc": f'{g.get("grain_type", "Wheat")} Grinding ({g["wheat_weight"]} kg)',
              "qty": f'{g["wheat_weight"]} kg', "rate": g.get("charge_per_kg", 0),
              "amount": g.get("total_charge", 0)}]
    method = normalise_method(g.get("payment_method"))
    if method == FLOUR_DEDUCTION and g.get("deducted_flour"):
        items.append({"desc": f'Less: {out_name} kept as grinding fee',
                      "qty": f'{g["deducted_flour"]} kg',
                      "rate": g.get("flour_unit_rate", 0),
                      "amount": -abs(g.get("flour_value", 0))})
        items.append({"desc": f'{out_name} delivered to customer',
                      "qty": f'{g.get("final_flour_delivered", g.get("customer_receives", 0))} kg',
                      "rate": "", "amount": ""})
    elif method == GRAIN_DEDUCTION and g.get("grain_qty"):
        items.append({"desc": f'Paid in {g.get("grain_item") or "grain"}',
                      "qty": f'{g["grain_qty"]} kg', "rate": "",
                      "amount": -abs(g.get("grain_value", 0))})
    return items

async def build_invoice_data(ref_id: str):
    sale = await db.sales.find_one({"id": ref_id})
    if sale:
        unit = await product_unit(sale.get("product_id"))
        return {"type": "Sale", "invoice_number": sale["invoice_number"], "date": sale["date"],
                "customer_name": sale["customer_name"], "payment_status": sale["payment_status"],
                "items": [{"desc": sale["product_name"], "qty": f'{sale["quantity"]} {unit}',
                           "rate": sale["price"], "amount": sale["total"]}], "total": sale["total"]}
    x = await db.exchanges.find_one({"id": ref_id})
    if x:
        items = [{"desc": "Wheat received", "qty": f'{x.get("wheat_qty", 0)} kg', "rate": "", "amount": ""},
                 {"desc": "Flour produced", "qty": f'{x.get("flour_produced", x.get("atta_given", 0))} kg', "rate": "", "amount": ""},
                 {"desc": "Grinding charge", "qty": f'{x.get("wheat_qty", 0)} kg',
                  "rate": x.get("grinding_rate", 0), "amount": x.get("grinding_charge", 0)}]
        if normalise_method(x.get("payment_method")) == FLOUR_DEDUCTION and x.get("deducted_flour"):
            items.append({"desc": "Less: flour deducted as grinding charges",
                          "qty": f'{x["deducted_flour"]} kg', "rate": x.get("flour_unit_rate", 0),
                          "amount": -abs(x.get("flour_value", 0))})
        items.append({"desc": "Final flour delivered",
                      "qty": f'{x.get("final_flour_delivered", x.get("atta_given", 0))} kg', "rate": "", "amount": ""})
        return {"type": "Exchange", "invoice_number": x.get("invoice_number") or "-", "date": x.get("date"),
                "customer_name": x.get("customer_name"), "payment_status": x.get("payment_status", "Paid"),
                "payment_method": normalise_method(x.get("payment_method")),
                "items": items, "total": x.get("grinding_charge", 0)}

    g = await db.grinding.find_one({"id": ref_id})
    if g:
        return {"type": "Grinding Service", "invoice_number": g["invoice_number"], "date": g["date"],
                "customer_name": g["customer_name"], "payment_status": g["payment_status"],
                "items": grinding_invoice_items(g), "total": g["total_charge"],
                "payment_method": normalise_method(g.get("payment_method"))}
    o = await db.oil.find_one({"id": ref_id})
    if o:
        items = [{"desc": f'{o["seed_type"]} Oil Extraction ({o["oil_extracted"]} L extracted)',
                  "qty": f'{o["quantity_received"]} kg', "rate": o["charge"], "amount": o["charge"]}]
        # Show the cake purchase as its own negative line, otherwise the
        # customer cannot see why the bill is lower than the extraction charge.
        if o.get("cake_sold_to_shop", 0):
            items.append({"desc": "Less: mustard cake purchased from customer",
                          "qty": f'{o["cake_sold_to_shop"]} kg',
                          "rate": o.get("cake_rate", 0),
                          "amount": -o.get("cake_value", 0)})
        return {"type": "Oil Extraction Service", "invoice_number": o["invoice_number"], "date": o["date"],
                "customer_name": o["customer_name"], "payment_status": o["payment_status"],
                "items": items,
                "total": o.get("total", o["charge"])}
    return None

@api_router.get("/invoices/{ref_id}/pdf")
async def invoice_pdf(ref_id: str, request: Request):
    await get_current_user(request)
    data = await build_invoice_data(ref_id)
    if not data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#B8860B"))
    elements = [Paragraph("Gangotri Flour &amp; Oil Mill", title),
                Paragraph("Wheat Grinding &amp; Oil Extraction Services", styles["Normal"]),
                Spacer(1, 12),
                Paragraph(f"<b>Invoice:</b> {data['invoice_number']} &nbsp;&nbsp; <b>Type:</b> {data['type']}", styles["Normal"]),
                Paragraph(f"<b>Date:</b> {data['date']} &nbsp;&nbsp; <b>Status:</b> {data['payment_status']}", styles["Normal"]),
                Paragraph(f"<b>Customer:</b> {data['customer_name']}", styles["Normal"]),
                Spacer(1, 16)]
    rows = [["Description", "Quantity", "Rate (Rs)", "Amount (Rs)"]]
    for it in data["items"]:
        def cell(v):
            return f'{v:.2f}' if isinstance(v, (int, float)) else (v or "")
        rows.append([it["desc"], it["qty"], cell(it["rate"]), cell(it["amount"])])
    rows.append(["", "", "Total", f'Rs {data["total"]:.2f}'])
    tbl = Table(rows, colWidths=[80*mm, 30*mm, 35*mm, 35*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B8860B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F5EEDC")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Thank you for your business!", styles["Italic"]))
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{data["invoice_number"]}.pdf"'})

# ---------------- Reports / Export ----------------

@api_router.get("/dashboard")
async def dashboard(user: dict = Depends(get_current_user)):
    products = await db.products.find().to_list(1000)
    sales = await db.sales.find().to_list(5000)
    purchases = await db.purchases.find().to_list(5000)
    grinding = await db.grinding.find().to_list(5000)
    oil = await db.oil.find().to_list(5000)
    expenses = await db.expenses.find().to_list(5000)

    today = datetime.now().strftime("%Y-%m-%d")
    month = datetime.now().strftime("%Y-%m")

    total_sales = sum(s.get("total", 0) for s in sales)
    total_purchases = sum(p.get("total", 0) for p in purchases)
    grinding_income = sum(g.get("total_charge", 0) for g in grinding)
    oil_income = sum(o.get("total", 0) for o in oil)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    service_income = grinding_income + oil_income
    total_income = total_sales + service_income

    # Profit charges the cost of what was actually sold, not everything bought.
    # Buying stock converts cash into inventory; it becomes a cost only when
    # that stock is sold. Subtracting whole purchases showed a loss whenever the
    # shop restocked — 250 kg bought against 25 kg sold read as 225 kg of loss.
    cost_by_name = {p.get("name"): p.get("cost_per_unit", 0) or 0 for p in products}
    total_cogs = sum(sale_cogs(s, cost_by_name) for s in sales)
    profit = total_income - total_cogs - total_expenses

    def day_sum(items, field):
        return sum(i.get(field, 0) for i in items if str(i.get("date", "")).startswith(today))
    def month_sum(items, field):
        return sum(i.get(field, 0) for i in items if str(i.get("date", "")).startswith(month))

    daily_income = day_sum(sales, "total") + day_sum(grinding, "total_charge") + day_sum(oil, "total")
    monthly_income = month_sum(sales, "total") + month_sum(grinding, "total_charge") + month_sum(oil, "total")

    # Outstanding is the unpaid balance, so a part-paid bill contributes only
    # what is still owed. Records written before part payment existed have no
    # balance_due, so fall back to the total when the status is not Paid.
    def owed(d, field):
        # A negative settlement is money the shop paid out, not a receivable.
        # max(...,0) keeps it out of Pending in both directions: it neither adds
        # to the figure nor quietly cancels a genuine debt from another bill.
        if (d.get(field, 0) or 0) < 0:
            return 0
        if d.get("balance_due") is not None:
            return max(d.get("balance_due", 0), 0)
        return 0 if d.get("payment_status") == "Paid" else max(d.get(field, 0), 0)

    pending_customer = 0.0
    for coll, field in [(sales, "total"), (grinding, "total_charge"), (oil, "total")]:
        pending_customer += sum(owed(d, field) for d in coll)
    supplier_dues = sum(owed(p, "total") for p in purchases)
    paid_to_customers = round(sum(abs(d.get("total", 0)) for d in oil if (d.get("total", 0) or 0) < 0), 2)

    low_stock = [{"name": p["name"], "stock": p.get("current_stock", 0), "threshold": p.get("low_stock_threshold", 0), "unit": p.get("unit", "kg")}
                 for p in products if p.get("current_stock", 0) <= p.get("low_stock_threshold", 0)]

    # last 6 months income vs expense
    trend = []
    for i in range(5, -1, -1):
        d = (datetime.now().replace(day=1) - timedelta(days=i * 30))
        m = d.strftime("%Y-%m")
        label = d.strftime("%b")
        inc = month_income_for(sales, "total", m) + month_income_for(grinding, "total_charge", m) + month_income_for(oil, "total", m)
        # Cost of goods sold that month, not stock bought that month, so the
        # line tracks the same profit the headline figure reports.
        month_cogs = sum(sale_cogs(s, cost_by_name) for s in sales if str(s.get("date", "")).startswith(m))
        exp = month_income_for(expenses, "amount", m) + month_cogs
        trend.append({"month": label, "income": round(inc, 2), "expense": round(exp, 2)})

    return {
        "total_income": round(total_income, 2),
        "total_sales": round(total_sales, 2),
        "total_purchases": round(total_purchases, 2),
        "service_income": round(service_income, 2),
        "grinding_orders": len(grinding),
        "oil_orders": len(oil),
        "total_expenses": round(total_expenses, 2),
        "total_cogs": round(total_cogs, 2),
        "profit": round(profit, 2),
        "daily_income": round(daily_income, 2),
        "monthly_income": round(monthly_income, 2),
        "pending_customer": round(pending_customer, 2),
        "supplier_dues": round(supplier_dues, 2),
        # Money handed back to customers when a by-product outweighed the charge.
        "paid_to_customers": paid_to_customers,
        "inventory_count": len(products),
        "total_stock": round(sum(p.get("current_stock", 0) for p in products), 2),
        "low_stock": low_stock,
        "trend": trend,
        "revenue_breakdown": [
            {"name": "Product Sales", "value": round(total_sales, 2)},
            {"name": "Grinding", "value": round(grinding_income, 2)},
            {"name": "Oil Extraction", "value": round(oil_income, 2)},
        ],
    }

def month_income_for(items, field, m):
    return sum(i.get(field, 0) for i in items if str(i.get("date", "")).startswith(m))

@api_router.get("/notifications")
async def notifications(user: dict = Depends(get_current_user)):
    notes = []
    for p in await db.products.find().to_list(1000):
        if p.get("current_stock", 0) <= p.get("low_stock_threshold", 0):
            notes.append({"type": "low_stock", "level": "warning",
                          "message": f'Low stock: {p["name"]} ({p.get("current_stock",0)} {p.get("unit","kg")} left)'})
    for coll, field, label in [("sales", "total", "sale"), ("grinding", "total_charge", "grinding"), ("oil", "total", "oil extraction")]:
        docs = await db[coll].find({"payment_status": {"$in": ["Pending", "Partial"]}}).to_list(1000)
        docs = [d for d in docs if (d.get(field, 0) or 0) > 0]
        for d in docs:
            due = d.get("balance_due", d.get(field, 0))
            part = " (part paid)" if d.get("payment_status") == "Partial" else ""
            notes.append({"type": "pending_payment", "level": "info",
                          "message": f'Pending payment from {d.get("customer_name","?")} - Rs {due:.0f}{part} ({label})'})
    for p in await db.purchases.find({"payment_status": {"$in": ["Pending", "Partial"]}}).to_list(1000):
        due = p.get("balance_due", p.get("total", 0))
        notes.append({"type": "supplier_due", "level": "info",
                      "message": f'Supplier due: {p.get("supplier_name","?")} - Rs {due:.0f}'})
    today_str = datetime.now().strftime("%Y-%m-%d")
    soon = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    for m in await db.maintenance.find().to_list(1000):
        nd = m.get("next_due_date", "")
        if nd and nd <= soon:
            overdue = nd < today_str
            notes.append({"type": "maintenance", "level": "warning" if overdue else "info",
                          "message": f'{"OVERDUE" if overdue else "Upcoming"} maintenance: {m.get("machine","?")} — {m.get("task","service")} (due {nd})'})
    return notes

@api_router.get("/export/{kind}")
async def export_excel(kind: str, request: Request):
    await get_current_user(request)
    coll_map = {"sales": "sales", "purchases": "purchases", "grinding": "grinding",
                "oil": "oil", "expenses": "expenses"}
    if kind not in coll_map:
        raise HTTPException(status_code=404, detail="Unknown report")
    docs = [clean(d) for d in await db[coll_map[kind]].find().to_list(5000)]
    wb = Workbook()
    ws = wb.active
    ws.title = kind
    if docs:
        headers = [k for k in docs[0].keys() if k not in ("created_at",)]
        ws.append(headers)
        for d in docs:
            ws.append([d.get(h, "") for h in headers])
    else:
        ws.append(["No data"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{kind}_report.xlsx"'})

# ---------------- Startup ----------------

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.login_attempts.create_index("identifier")
    await db.password_resets.create_index("token_hash", unique=True)
    # Mongo drops expired reset records on its own; the endpoint also checks.
    await db.password_resets.create_index("expires_at", expireAfterSeconds=0)
    await db.reset_attempts.create_index("identifier")
    # One automatic posting per source, enforced by the database rather than by
    # every caller remembering to check first.
    # Partial rather than sparse: sparse only skips documents missing the field,
    # so any row storing an explicit null would still collide with the others.
    await db.bank_txns.create_index("source_ref", unique=True,
                                    partialFilterExpression={"source_ref": {"$type": "string"}})
    await db.bank_txns.create_index([("bank_id", 1), ("date", -1)])
    await db.bank_txns.create_index("reconciled")
    await db.bank_accounts.create_index("account_digits", sparse=True)
    # The ledger and search paths scan these on every customer and supplier row;
    # unindexed they are full collection scans that grow with total sales.
    for coll, field in (("sales", "customer_name"), ("grinding", "customer_name"),
                        ("oil", "customer_name"), ("purchases", "supplier_name"),
                        ("payments", "party_name")):
        await db[coll].create_index(field)
    for coll in ("sales", "purchases", "grinding", "oil", "expenses", "payments"):
        await db[coll].create_index("date")
    for coll in ("sales", "grinding", "oil", "exchanges"):
        await db[coll].create_index("invoice_number")
    for coll in ("customers", "suppliers"):
        await db[coll].create_index("phone")
    await db.bank_txns.create_index("mode")
    await consolidate_parties()
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@agrimill.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({"email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Admin", "role": "admin", "created_at": now_iso()})
    elif not existing.get("password_self_managed") and not verify_password(admin_password, existing["password_hash"]):
        # Re-sync from ADMIN_PASSWORD only while the admin has never changed it
        # themselves. Without this guard, startup runs on every wake from idle
        # and would silently revert a password set via the reset flow.
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
    # No demo staff account is seeded. Startup runs on every wake from idle, so
    # a seeded account would keep reappearing after being deleted. Admins add
    # staff from Settings instead.
    await consolidate_products()
    await seed_products()
    await get_settings_doc()

# ==================== Mill Production & Advanced Logic ====================

DEFAULT_PRODUCTS = [
    {"name": "Wheat Crop", "category": "Wheat Crop", "unit": "kg", "low_stock_threshold": 100},
    {"name": "Atta", "category": "Flour", "unit": "kg", "low_stock_threshold": 50},
    {"name": "Multigrain Atta", "category": "Flour", "unit": "kg", "low_stock_threshold": 30},
    {"name": "Besan", "category": "Flour", "unit": "kg", "low_stock_threshold": 30},
    {"name": "Makka Atta", "category": "Flour", "unit": "kg", "low_stock_threshold": 30},
    {"name": "Bajra Atta", "category": "Flour", "unit": "kg", "low_stock_threshold": 30},
    {"name": "Sattu", "category": "Flour", "unit": "kg", "low_stock_threshold": 20},
    {"name": "Wheat Bran", "category": "Bran", "unit": "kg", "low_stock_threshold": 30},
    # Raw inputs for the Besan and Masala mills. Production resolves a mill's
    # input and outputs by product name, so a missing row here makes that mill
    # unusable with no visible reason.
    {"name": "Gram (Chana)", "category": "Other", "unit": "kg", "low_stock_threshold": 50},
    {"name": "Whole Spices", "category": "Other", "unit": "kg", "low_stock_threshold": 20},
    {"name": "Masala", "category": "Masala", "unit": "kg", "low_stock_threshold": 20},
    {"name": "Mustard Seeds", "category": "Oil Seeds", "unit": "kg", "low_stock_threshold": 100},
    {"name": "Mustard Oil", "category": "Edible Oil", "unit": "litre", "low_stock_threshold": 20},
    {"name": "Mustard Oil Cake", "category": "Oil Cake", "unit": "kg", "low_stock_threshold": 30},
    {"name": "Packing Bags", "category": "Packing", "unit": "pcs", "low_stock_threshold": 100},
]

async def consolidate_products():
    """Backfill name_key and fold pre-existing duplicate products into one row.

    Products created before name_key existed can share a name. That breaks
    adjust_stock_by_name, which would update an arbitrary one of them, and it
    blocks the unique index below. Runs on every boot but is a no-op once clean.
    """
    async for p in db.products.find({"name_key": {"$exists": False}}):
        await db.products.update_one({"_id": p["_id"]}, {"$set": {"name_key": product_key(p.get("name", ""))}})

    groups = await db.products.aggregate([
        {"$group": {"_id": "$name_key", "ids": {"$push": "$id"}, "n": {"$sum": 1}}},
        {"$match": {"n": {"$gt": 1}}},
    ]).to_list(1000)

    for g in groups:
        dupes = await db.products.find({"id": {"$in": g["ids"]}}).to_list(100)
        # Keep the row carrying the most stock; it is the one in real use.
        dupes.sort(key=lambda d: (d.get("current_stock", 0) or 0, d.get("created_at") or ""), reverse=True)
        keeper, rest = dupes[0], dupes[1:]

        total_stock = sum(d.get("current_stock", 0) or 0 for d in dupes)
        # Weighted average so the merged cost basis still reflects what was paid.
        valued = sum((d.get("current_stock", 0) or 0) * (d.get("cost_per_unit", 0) or 0) for d in dupes)
        merged = {
            "current_stock": round(total_stock, 3),
            "cost_per_unit": round(valued / total_stock, 4) if total_stock > 0 else 0,
            "rate": max((d.get("rate", 0) or 0) for d in dupes),
            "low_stock_threshold": max((d.get("low_stock_threshold", 0) or 0) for d in dupes),
        }
        await db.products.update_one({"id": keeper["id"]}, {"$set": merged})

        # Repoint history at the survivor so unit lookups and edits keep working.
        for d in rest:
            for coll, field in ((db.sales, "product_id"), (db.purchases, "product_id"),
                                (db.production, "input_product_id"), (db.production, "outputs.$[o].product_id")):
                if field.startswith("outputs"):
                    await coll.update_many({"outputs.product_id": d["id"]},
                                           {"$set": {"outputs.$[o].product_id": keeper["id"]}},
                                           array_filters=[{"o.product_id": d["id"]}])
                else:
                    await coll.update_many({field: d["id"]}, {"$set": {field: keeper["id"]}})
            await db.products.delete_one({"id": d["id"]})

        logger.warning("Merged %d duplicate product rows named %r into one (stock %s)",
                       len(rest) + 1, keeper.get("name"), merged["current_stock"])

    await db.products.create_index("name_key", unique=True)

async def consolidate_parties():
    """Give every party a normalised key and fold same-name duplicates together.

    Parties created before the key existed can differ only by casing or spacing,
    which reads as two ledgers for one person. The survivor keeps the richest
    record rather than whichever happened to be first.
    """
    for coll in ("customers", "suppliers"):
        async for row in db[coll].find({"party_key": {"$exists": False}}):
            await db[coll].update_one({"_id": row["_id"]},
                                      {"$set": {"party_key": party_key(row.get("name", ""))}})

        groups = await db[coll].aggregate([
            {"$group": {"_id": "$party_key", "ids": {"$push": "$id"}, "n": {"$sum": 1}}},
            {"$match": {"n": {"$gt": 1}}},
        ]).to_list(1000)
        for g in groups:
            rows = await db[coll].find({"id": {"$in": g["ids"]}}).to_list(100)
            # Most filled-in record wins; a bare duplicate should not overwrite
            # a party someone took the time to enter properly.
            rows.sort(key=lambda r: sum(1 for f in ("phone", "address", "gstin", "pan_aadhaar") if r.get(f)),
                      reverse=True)
            keeper, rest = rows[0], rows[1:]
            merged = {f: keeper.get(f) or next((r.get(f) for r in rest if r.get(f)), "")
                      for f in ("phone", "address", "gstin", "pan_aadhaar")}
            merged["opening_balance"] = sum(r.get("opening_balance", 0) or 0 for r in rows)
            merged["credit_limit"] = max((r.get("credit_limit", 0) or 0) for r in rows)
            await db[coll].update_one({"id": keeper["id"]}, {"$set": merged})
            for r in rest:
                await db[coll].delete_one({"id": r["id"]})
            logger.warning("Merged %d duplicate %s records named %r", len(rows), coll, keeper.get("name"))

        await db[coll].create_index("party_key", unique=True,
                                    partialFilterExpression={"party_key": {"$type": "string"}})

async def seed_products():
    """Populate the catalogue once, on a brand new database only.

    This used to top the catalogue up on every boot, re-inserting any default
    product that was missing. Startup runs again each time Render's free tier
    wakes from idle, so a product the mill deleted reappeared within the hour as
    a blank row — and it looked random, because it came back on the next wake
    rather than straight after the delete. Deleting is the operator's decision;
    seeding is only for giving an empty database somewhere to start.
    """
    settings = await db.settings.find_one({"id": "config"}) or {}
    if settings.get("catalogue_seeded"):
        return

    # An existing database already has whatever catalogue the mill wants,
    # including the deliberate absences. Adopt it rather than topping it up.
    if await db.products.count_documents({}) == 0:
        for p in DEFAULT_PRODUCTS:
            await db.products.insert_one({"id": str(uuid.uuid4()), **p, "name_key": product_key(p["name"]),
                "current_stock": 0, "rate": 0, "cost_per_unit": 0, "created_at": now_iso()})

    await db.settings.update_one({"id": "config"}, {"$set": {"catalogue_seeded": True}}, upsert=True)

# What the mill will grind, and what each item yields. The operator can add to
# this from the grinding form — masala, a new millet — without a code change.
DEFAULT_GRAIN_TYPES = [
    {"name": "Wheat", "output": "Atta"},
    {"name": "Gram (Chana)", "output": "Besan"},
    {"name": "Multigrain Mix", "output": "Multigrain Atta"},
    {"name": "Maize (Makka)", "output": "Makka Atta"},
    {"name": "Bajra", "output": "Bajra Atta"},
    {"name": "Roasted Gram", "output": "Sattu"},
]

async def get_settings_doc():
    s = await db.settings.find_one({"id": "config"})
    if not s:
        s = {"id": "config", "washed_loss": 2.5, "unwashed_loss": 5.0, "starting_cash": 0,
             "grinding_rate": 2.0, "flour_deduction_percent": 5.0, "flour_rate": 0}
        await db.settings.insert_one(s)
    defaults = {"grinding_rate": 2.0, "flour_deduction_percent": 5.0, "flour_rate": 0}
    missing = {k: v for k, v in defaults.items() if s.get(k) is None}
    if missing:
        await db.settings.update_one({"id": "config"}, {"$set": missing})
        s.update(missing)
    if not s.get("grain_types"):
        await db.settings.update_one({"id": "config"}, {"$set": {"grain_types": DEFAULT_GRAIN_TYPES}})
        s["grain_types"] = DEFAULT_GRAIN_TYPES
    return clean(s)

class GrainTypeBody(BaseModel):
    name: str
    output: str

@api_router.get("/grain-types")
async def list_grain_types(user: dict = Depends(get_current_user)):
    return (await get_settings_doc()).get("grain_types", DEFAULT_GRAIN_TYPES)

@api_router.post("/grain-types")
async def add_grain_type(body: GrainTypeBody, user: dict = Depends(get_current_user)):
    name = body.name.strip()
    output = (body.output or "").strip() or f"{name} Atta"
    if not name:
        raise HTTPException(status_code=400, detail="Enter a name")
    types = (await get_settings_doc()).get("grain_types", [])
    if any(t.get("name", "").lower() == name.lower() for t in types):
        raise HTTPException(status_code=400, detail=f'"{name}" is already in the list')
    types.append({"name": name, "output": output})
    await db.settings.update_one({"id": "config"}, {"$set": {"grain_types": types}})
    # Ground output has to exist in the catalogue for the shop's cut to be stocked.
    if await db.products.find_one({"name_key": product_key(output)}) is None:
        await db.products.insert_one({"id": str(uuid.uuid4()), "name": output, "name_key": product_key(output),
            "category": "Flour", "unit": "kg", "current_stock": 0, "cost_per_unit": 0, "rate": 0,
            "low_stock_threshold": 20, "created_at": now_iso()})
    await log_audit(user, "Added grinding item", f"{name} → {output}")
    return types

async def adjust_stock_by_name(name, delta):
    """Move stock for a product identified by name.

    Matches the normalised key so casing and stray spaces cannot miss the row.
    Names are unique (see product_key), so this always targets exactly one.
    """
    res = await db.products.update_one({"name_key": product_key(name)},
                                       {"$inc": {"current_stock": round(delta, 3)}})
    if res.matched_count == 0:
        logger.warning("adjust_stock_by_name: no product named %r; %+g not applied", name, delta)
        return
    # Re-round after the increment: repeated $inc on floats accumulates binary
    # error, so a figure that should read 506.8 drifts to 506.79999999999995.
    prod = await db.products.find_one({"name_key": product_key(name)})
    if prod is not None:
        await db.products.update_one({"name_key": product_key(name)},
                                     {"$set": {"current_stock": round(prod.get("current_stock", 0) or 0, 3)}})

async def add_stock_by_name_with_cost(name, qty, total_cost):
    """Move stock by name while folding the money paid into the cost basis.

    adjust_stock_by_name only touches quantity, which is right for stock that
    arrived free. Stock the shop actually bought must go through here, or its
    cost_per_unit stays stale and sales-analytics reports the whole sale price
    as profit.
    """
    p = await db.products.find_one({"name_key": product_key(name)})
    if not p:
        logger.warning("add_stock_by_name_with_cost: no product named %r; %+g not applied", name, qty)
        return
    await add_stock_with_cost(p["id"], qty, total_cost)

async def add_stock_with_cost(pid, qty, total_cost):
    p = await db.products.find_one({"id": pid})
    if not p:
        # Returning quietly here is how a purchase could be recorded against a
        # stale id while the shelf figure never moved. Callers resolve the
        # product first; this is the last line of defence and must be visible.
        logger.warning("Stock move of %s skipped: no product with id %r", qty, pid)
        return
    old_stock = p.get("current_stock", 0)
    old_cost = p.get("cost_per_unit", 0)
    new_stock = old_stock + qty
    new_cost = ((old_stock * old_cost) + total_cost) / new_stock if new_stock > 0 else 0
    await db.products.update_one({"id": pid}, {"$set": {
        "current_stock": round(new_stock, 3), "cost_per_unit": round(new_cost, 4)}})

# ---- build helpers ----
# ---------------- Paying in kind ----------------
# Many customers settle grinding by leaving flour or grain behind rather than
# handing over cash. The charge is still a charge: it is computed and shown
# whichever way it is paid, so the mill can see what its grinding actually
# earned. Only the settlement differs.

MONEY_METHODS = ("Cash", "UPI", "Bank", "NEFT", "RTGS", "IMPS", "Cheque")
FLOUR_DEDUCTION = "Flour Deduction"
GRAIN_DEDUCTION = "Grain Deduction"
KIND_METHODS = (FLOUR_DEDUCTION, GRAIN_DEDUCTION)
GRINDING_METHODS = MONEY_METHODS + KIND_METHODS

# How much flour to keep back: a share of the output, a flat weight, or however
# much covers the charge at the flour rate.
DEDUCTION_BASES = ("Percent", "Weight", "Value")

def normalise_method(method: str) -> str:
    """Map what a record stored onto today's vocabulary.

    "Grain" was the old single in-kind option and always meant flour kept back,
    so it reads as a flour deduction rather than falling through to Cash and
    misreporting an in-kind job as money taken.
    """
    m = (method or "Cash").strip()
    if m in GRINDING_METHODS:
        return m
    if m.lower() in ("grain", "grain/material", "material"):
        return FLOUR_DEDUCTION
    return clean_mode(m) if clean_mode(m) in MONEY_METHODS else "Cash"

def settled_in_kind(method: str) -> bool:
    return normalise_method(method) in KIND_METHODS

async def flour_unit_rate(product_name: str) -> float:
    """What a kilo of kept-back flour is valued at."""
    settings = await get_settings_doc()
    configured = settings.get("flour_rate", 0) or 0
    if configured > 0:
        return round(configured, 4)
    prod = await db.products.find_one({"name_key": product_key(product_name or "Atta")})
    return round((prod or {}).get("rate", 0) or 0, 4)

async def compute_flour_deduction(*, output_qty, total_charge, basis, percent, weight, unit_rate):
    """Flour kept back as the grinding fee, and what it is worth.

    Value basis keeps exactly enough flour to cover the charge, which is what a
    mill quoting in rupees actually does. Percent and Weight let the operator
    work the way their customers expect instead.
    """
    b = basis if basis in DEDUCTION_BASES else "Percent"
    if b == "Weight":
        deducted = max(float(weight or 0), 0)
    elif b == "Value":
        deducted = round(total_charge / unit_rate, 3) if unit_rate > 0 else 0.0
    else:
        deducted = round(output_qty * max(float(percent or 0), 0) / 100, 3)
    deducted = round(min(deducted, output_qty), 3)
    return deducted, round(deducted * unit_rate, 2)

async def post_material_ledger(*, date, party_name, item, qty, unit, value, direction,
                               ref_id=None, source=None, note=""):
    """Record stock moving in settlement of a bill.

    Kept apart from the cash and bank ledgers on purpose: this is grain and
    flour changing hands, not money, and folding it into either would overstate
    what the mill actually took in rupees.
    """
    if not qty:
        return
    existing = await db.material_ledger.find_one({"ref_id": ref_id, "source": source}) if ref_id else None
    doc = {"date": date, "party_name": party_name, "item": item, "qty": round(qty, 3),
           "unit": unit, "value": round(value or 0, 2), "direction": direction,
           "ref_id": ref_id, "source": source, "note": note}
    if existing:
        await db.material_ledger.update_one({"id": existing["id"]}, {"$set": doc})
        return existing["id"]
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    await db.material_ledger.insert_one(doc)
    return doc["id"]

async def unpost_material_ledger(ref_id: str):
    if ref_id:
        await db.material_ledger.delete_many({"ref_id": ref_id})

@api_router.get("/material-ledger")
async def material_ledger(start: Optional[str] = None, end: Optional[str] = None,
                          party: Optional[str] = None, item: Optional[str] = None,
                          user: dict = Depends(get_current_user)):
    q = {}
    if party: q["party_name"] = party
    if item: q["item"] = item
    if start or end:
        rng = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end
        q["date"] = rng
    rows = [clean(r) for r in await db.material_ledger.find(q).sort("date", -1).to_list(20000)]
    return {"rows": rows,
            "total_in": round(sum(r["qty"] for r in rows if r.get("direction") == "in"), 3),
            "total_value": round(sum(r.get("value", 0) for r in rows if r.get("direction") == "in"), 2)}

async def build_grinding_doc(body: GrindingBody):
    inv = await next_invoice_number()
    output_atta = round(body.wheat_weight * (1 - body.loss_percent / 100), 2)
    loss_kg = round(body.wheat_weight - output_atta, 2)
    method = normalise_method(body.payment_method)
    output_name = body.output_product or "Atta"

    # The charge is what the grinding was worth and is computed the same way
    # every time. Previously paying in kind set it to zero, which hid the
    # mill's own earnings whenever a customer settled with flour.
    total_charge = round(body.wheat_weight * (body.charge_per_kg or 0), 2)

    deducted_flour = 0.0
    flour_value = 0.0
    unit_rate = 0.0
    grain_qty = round(body.grain_qty or 0, 3)
    grain_value = round(body.grain_value or 0, 2)

    if method == FLOUR_DEDUCTION:
        unit_rate = await flour_unit_rate(output_name)
        settings = await get_settings_doc()
        percent = body.deduction_percent if body.deduction_percent is not None else settings.get("flour_deduction_percent", 5)
        # grain_fee_kg is what older records used for the same idea. A stated
        # weight is an instruction, so honour it rather than overwriting it with
        # the configured percentage.
        weight = body.deduction_weight if body.deduction_weight is not None else body.grain_fee_kg
        basis = body.deduction_basis
        if body.deduction_weight is None and body.deduction_basis == "Percent" and body.grain_fee_kg:
            basis = "Weight"
        deducted_flour, flour_value = await compute_flour_deduction(
            output_qty=output_atta, total_charge=total_charge, basis=basis,
            percent=percent, weight=weight, unit_rate=unit_rate)
    elif method == GRAIN_DEDUCTION:
        grain_qty = round(body.grain_qty or body.grain_fee_kg or 0, 3)
        grain_value = round(body.grain_value or 0, 2)

    customer_receives = round(output_atta - deducted_flour, 3)

    d = body.model_dump()
    d.update({"id": str(uuid.uuid4()), "invoice_number": inv, "output_atta": output_atta,
              "loss_kg": loss_kg, "customer_receives": customer_receives,
              "payment_method": method, "total_charge": total_charge,
              "deducted_flour": deducted_flour, "flour_unit_rate": unit_rate,
              "flour_value": flour_value, "final_flour_delivered": customer_receives,
              "grain_qty": grain_qty, "grain_value": grain_value,
              # Kept in step so old readers of grain_fee_kg still see the weight.
              "grain_fee_kg": deducted_flour if method == FLOUR_DEDUCTION else grain_qty,
              "created_at": now_iso()})
    return d

def grinding_output_name(doc: dict) -> str:
    """Product the shop's grain fee is stocked as. Older rows are all atta."""
    return doc.get("output_product") or "Atta"

async def build_oil_doc(body: OilBody):
    inv = await next_invoice_number()
    allocated = round(body.retained_cake + body.cake_sold_to_shop, 3)
    if allocated > body.oil_cake_produced + 0.009:
        raise HTTPException(status_code=400, detail=(
            f"Cake kept plus cake bought is {allocated} kg but only "
            f"{body.oil_cake_produced} kg was produced"))
    if body.cake_sold_to_shop and body.cake_rate <= 0:
        raise HTTPException(status_code=400, detail="Enter a rate for the cake being bought")
    # The shop pays for the cake by knocking its value off the extraction charge,
    # so the net is what the customer actually owes. It may go negative when the
    # cake is worth more than the grinding; that is a genuine credit to the
    # customer and flows through the ledger as one.
    cake_value = round(body.cake_sold_to_shop * body.cake_rate, 2)
    d = body.model_dump()
    d.update({"id": str(uuid.uuid4()), "invoice_number": inv,
              "cake_value": cake_value,
              "total": round(body.charge - cake_value, 2),
              "customer_oil": round(body.oil_extracted - body.retained_oil, 2),
              "customer_cake": round(body.oil_cake_produced - body.retained_cake - body.cake_sold_to_shop, 2),
              "created_at": now_iso()})
    return d

async def apply_grinding_effects(doc, sign):
    method = normalise_method(doc.get("payment_method"))
    if method == FLOUR_DEDUCTION:
        kept = doc.get("deducted_flour", doc.get("grain_fee_kg", 0)) or 0
        if kept:
            await adjust_stock_by_name(grinding_output_name(doc), sign * kept)
            if sign > 0:
                await post_material_ledger(
                    date=doc.get("date"), party_name=doc.get("customer_name", ""),
                    item=grinding_output_name(doc), qty=kept, unit="kg",
                    value=doc.get("flour_value", 0), direction="in",
                    ref_id=doc["id"], source="grinding",
                    note=f'Grinding fee for invoice {doc.get("invoice_number", "")}')
            else:
                await unpost_material_ledger(doc["id"])
    elif method == GRAIN_DEDUCTION:
        qty = doc.get("grain_qty", 0) or 0
        item = doc.get("grain_item") or doc.get("grain_type") or "Wheat Crop"
        if qty:
            await adjust_stock_by_name(item, sign * qty)
            if sign > 0:
                await post_material_ledger(
                    date=doc.get("date"), party_name=doc.get("customer_name", ""),
                    item=item, qty=qty, unit="kg", value=doc.get("grain_value", 0),
                    direction="in", ref_id=doc["id"], source="grinding",
                    note=f'Grinding paid in grain, invoice {doc.get("invoice_number", "")}')
            else:
                await unpost_material_ledger(doc["id"])

async def apply_oil_effects(doc, sign):
    if doc.get("retained_oil", 0):
        await adjust_stock_by_name("Mustard Oil", sign * doc["retained_oil"])
    if doc.get("retained_cake", 0):
        # Kept as the processing fee: quantity only, nothing was paid for it.
        await adjust_stock_by_name("Mustard Oil Cake", sign * doc["retained_cake"])
    if doc.get("cake_sold_to_shop", 0):
        # Bought from the customer: the value must enter the weighted-average
        # cost basis, and sign flips both quantity and cost together so an edit
        # or delete reverses it exactly.
        await add_stock_by_name_with_cost("Mustard Oil Cake",
                                          sign * doc["cake_sold_to_shop"],
                                          sign * doc.get("cake_value", 0))

# ---- Settings ----
class SettingsBody(BaseModel):
    washed_loss: float
    unwashed_loss: float
    starting_cash: float = 0
    # Default grinding charge. Every job still carries its own rate; this is
    # what a new job starts from.
    grinding_rate: float = 2.0
    # Share of the flour kept back when the customer pays in kind instead of
    # cash. 5, 10, 15 — whatever the mill works on.
    flour_deduction_percent: float = 5.0
    # What a kilo of kept-back flour is worth. Zero falls back to the product's
    # own rate, so the mill only sets this if it prices deductions differently.
    flour_rate: float = 0

@api_router.get("/settings")
async def read_settings(user: dict = Depends(get_current_user)):
    return await get_settings_doc()

@api_router.put("/settings")
async def write_settings(body: SettingsBody, user: dict = Depends(get_current_user)):
    await db.settings.update_one({"id": "config"}, {"$set": body.model_dump()}, upsert=True)
    return await get_settings_doc()

# ---- Production (converts input product into outputs) ----
class ProdOutput(BaseModel):
    product_id: str
    product_name: str
    quantity: float

class ProductionBody(BaseModel):
    date: str
    mill: str
    input_product_id: str
    input_product_name: str
    input_quantity: float
    outputs: List[ProdOutput]

@api_router.get("/production")
async def get_production(user: dict = Depends(get_current_user)):
    return [clean(p) for p in await db.production.find().sort("date", -1).to_list(2000)]

async def build_production_doc(body: ProductionBody) -> dict:
    """Cost the run and shape the stored record. Applies no stock changes."""
    inp = await db.products.find_one({"id": body.input_product_id})
    input_cost = round(body.input_quantity * (inp or {}).get("cost_per_unit", 0), 2)
    total_out = sum(o.quantity for o in body.outputs) or 1
    out_records = []
    for o in body.outputs:
        allocated = round(input_cost * (o.quantity / total_out), 2)
        # product_id is persisted so a later edit or delete can reverse the exact
        # rows this run touched, rather than guessing from the name.
        out_records.append({"product_id": o.product_id, "product_name": o.product_name,
                            "quantity": o.quantity, "cost": allocated,
                            "cost_per_unit": round(allocated / o.quantity, 3) if o.quantity else 0})
    return {"id": str(uuid.uuid4()), "date": body.date, "mill": body.mill,
            "input_product_id": body.input_product_id,
            "input_product_name": body.input_product_name, "input_quantity": body.input_quantity,
            "input_cost": input_cost, "outputs": out_records, "created_at": now_iso()}

async def apply_production_effects(doc: dict):
    """Consume the input and bank the outputs, carrying cost across."""
    if doc.get("input_product_id"):
        await add_stock_with_cost(doc["input_product_id"], -doc.get("input_quantity", 0), -doc.get("input_cost", 0))
    else:
        await adjust_stock_by_name(doc.get("input_product_name", ""), -doc.get("input_quantity", 0))
    for o in doc.get("outputs", []):
        if o.get("product_id"):
            await add_stock_with_cost(o["product_id"], o.get("quantity", 0), o.get("cost", 0))
        else:
            await adjust_stock_by_name(o.get("product_name", ""), o.get("quantity", 0))

@api_router.post("/production")
async def create_production(body: ProductionBody, user: dict = Depends(get_current_user)):
    inp = await db.products.find_one({"id": body.input_product_id})
    if not inp:
        raise HTTPException(status_code=404, detail="Input product not found")
    if body.input_quantity > inp.get("current_stock", 0):
        raise HTTPException(status_code=400, detail=f"Not enough stock (have {inp.get('current_stock',0)})")
    doc = await build_production_doc(body)
    await apply_production_effects(doc)
    await db.production.insert_one(doc)
    await log_audit(user, "Production run", f"{body.mill}: {body.input_quantity} {body.input_product_name} → " + ", ".join(f"{o.quantity} {o.product_name}" for o in body.outputs))
    return clean(doc)

async def reverse_production_effects(doc: dict):
    """Undo a production run's stock movements: return the input, remove the outputs."""
    if doc.get("input_product_id"):
        await add_stock_with_cost(doc["input_product_id"], doc.get("input_quantity", 0), doc.get("input_cost", 0))
    else:
        await adjust_stock_by_name(doc.get("input_product_name", ""), doc.get("input_quantity", 0))
    for o in doc.get("outputs", []):
        if o.get("product_id"):
            await add_stock_with_cost(o["product_id"], -o.get("quantity", 0), -o.get("cost", 0))
        else:
            await adjust_stock_by_name(o.get("product_name", ""), -o.get("quantity", 0))

@api_router.put("/production/{pid}")
async def edit_production(pid: str, body: ProductionBody, user: dict = Depends(get_current_user)):
    old = await db.production.find_one({"id": pid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    await reverse_production_effects(old)
    inp = await db.products.find_one({"id": body.input_product_id})
    if not inp:
        # Put the original movements back rather than leaving the run half-undone.
        await apply_production_effects(old)
        raise HTTPException(status_code=404, detail="Input product not found")
    if body.input_quantity > inp.get("current_stock", 0):
        await apply_production_effects(old)
        raise HTTPException(status_code=400, detail=f"Not enough stock (have {inp.get('current_stock',0)})")
    doc = await build_production_doc(body)
    doc["id"] = pid
    doc["created_at"] = old.get("created_at", now_iso())
    await apply_production_effects(doc)
    await db.production.replace_one({"id": pid}, doc)
    await log_audit(user, "Edited production run", f'{body.mill}: {body.input_quantity} {body.input_product_name}')
    return clean(doc)

@api_router.delete("/production/{pid}")
async def delete_production(pid: str, user: dict = Depends(require_admin)):
    doc = await db.production.find_one({"id": pid})
    if doc:
        # Previously this deleted the record but left its stock movements in
        # place, so the flour it produced stayed on the books forever.
        await reverse_production_effects(doc)
        await db.production.delete_one({"id": pid})
    return {"message": "deleted"}

# ---- Exchange (wheat crop for atta) ----
class ExchangeBody(BaseModel):
    date: str
    customer_name: str
    wheat_qty: float
    washed: bool = True
    loss_percent: float = 2.5
    atta_given: float
    # Grinding is a separate charge from the swap itself. It was previously
    # folded into the quantities and never shown, so the mill could not see
    # what its grinding earned on an exchange.
    grinding_rate: Optional[float] = None
    payment_method: str = "Cash"
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None
    deduction_basis: str = "Value"
    deduction_percent: Optional[float] = None
    deduction_weight: Optional[float] = None
    payment_status: str = "Paid"

async def build_exchange_doc(body: ExchangeBody, existing: dict = None) -> dict:
    """Price an exchange and work out what the customer actually walks out with.

    atta_given is the flour the swap produces. When grinding is paid in flour,
    the fee comes out of that, so the delivered figure is lower — and both are
    recorded, because the customer needs to see why.
    """
    settings = await get_settings_doc()
    rate = body.grinding_rate if body.grinding_rate is not None else (settings.get("grinding_rate", 2) or 0)
    method = normalise_method(body.payment_method)
    charge = round(body.wheat_qty * rate, 2)

    deducted = 0.0
    value = 0.0
    unit_rate = 0.0
    if method == FLOUR_DEDUCTION:
        unit_rate = await flour_unit_rate("Atta")
        percent = body.deduction_percent if body.deduction_percent is not None else settings.get("flour_deduction_percent", 5)
        deducted, value = await compute_flour_deduction(
            output_qty=body.atta_given, total_charge=charge, basis=body.deduction_basis,
            percent=percent, weight=body.deduction_weight, unit_rate=unit_rate)

    doc = {**(existing or {}), **body.model_dump(),
           "loss_kg": round(body.wheat_qty * body.loss_percent / 100, 2),
           "grinding_rate": round(rate, 2), "grinding_charge": charge,
           "payment_method": method,
           "flour_produced": round(body.atta_given, 3),
           "deducted_flour": deducted, "flour_unit_rate": unit_rate, "flour_value": value,
           "final_flour_delivered": round(body.atta_given - deducted, 3)}
    return doc

async def apply_exchange_effects(doc: dict, sign: int):
    """Wheat in, flour out, and any flour kept back as the grinding fee.

    The fee stays with the shop, so only what the customer actually takes comes
    off stock — the deduction is netted against the flour going out.
    """
    await adjust_stock_by_name("Wheat Crop", sign * doc.get("wheat_qty", 0))
    delivered = doc.get("final_flour_delivered", doc.get("atta_given", 0)) or 0
    await adjust_stock_by_name("Atta", -sign * delivered)
    if normalise_method(doc.get("payment_method")) == FLOUR_DEDUCTION and doc.get("deducted_flour"):
        if sign > 0:
            await post_material_ledger(
                date=doc.get("date"), party_name=doc.get("customer_name", ""),
                item="Atta", qty=doc["deducted_flour"], unit="kg",
                value=doc.get("flour_value", 0), direction="in",
                ref_id=doc["id"], source="exchange",
                note="Grinding fee kept as flour on exchange")
        else:
            await unpost_material_ledger(doc["id"])

async def settle_exchange(doc: dict, body: ExchangeBody):
    """Post the grinding charge to whichever ledger the payment names."""
    charge = doc.get("grinding_charge", 0) or 0
    if charge <= 0:
        return
    method = normalise_method(doc.get("payment_method"))
    if method == FLOUR_DEDUCTION:
        await add_credit("customer", doc["customer_name"], charge, doc["date"], doc["id"],
                         f'Exchange grinding · {doc.get("deducted_flour", 0)} kg flour kept', mode="Flour")
    elif doc.get("payment_status", "Paid") == "Paid":
        await add_credit("customer", doc["customer_name"], charge, doc["date"], doc["id"],
                         "Exchange grinding charge", mode=body.payment_mode or method, bank_id=body.bank_id)

@api_router.get("/exchanges")
async def get_exchanges(user: dict = Depends(get_current_user)):
    return [clean(e) for e in await db.exchanges.find().sort("date", -1).to_list(2000)]

@api_router.post("/exchanges")
async def create_exchange(body: ExchangeBody, user: dict = Depends(get_current_user)):
    doc = await build_exchange_doc(body)
    doc["id"] = str(uuid.uuid4())
    doc["invoice_number"] = await next_invoice_number()
    doc["created_at"] = now_iso()
    # Only the flour the customer takes leaves stock; anything kept as the fee
    # stays with the shop, so that is the figure to check against.
    atta = await db.products.find_one({"name_key": product_key("Atta")})
    needed = doc["final_flour_delivered"]
    if atta and needed > atta.get("current_stock", 0):
        raise HTTPException(status_code=400, detail=f"Not enough Atta stock (have {atta.get('current_stock',0)} kg)")
    await db.exchanges.insert_one(doc)
    await db.invoices.insert_one({"id": str(uuid.uuid4()), "invoice_number": doc["invoice_number"],
        "type": "Exchange", "ref_id": doc["id"], "customer_name": doc["customer_name"],
        "date": doc["date"], "total": doc["grinding_charge"],
        "payment_status": doc.get("payment_status", "Paid"), "created_at": now_iso()})
    await apply_exchange_effects(doc, 1)
    await settle_exchange(doc, body)
    await log_audit(user, "Exchange", f'{body.customer_name}: {body.wheat_qty} kg wheat, grinding Rs {doc["grinding_charge"]} by {doc["payment_method"]}')
    return clean(doc)

@api_router.put("/exchanges/{eid}")
async def edit_exchange(eid: str, body: ExchangeBody, user: dict = Depends(get_current_user)):
    old = await db.exchanges.find_one({"id": eid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    # Undo the original swap first, so the stock check below sees the atta that
    # this very exchange had taken out.
    await apply_exchange_effects(old, -1)
    doc = await build_exchange_doc(body, existing=old)
    doc["id"] = eid
    atta = await db.products.find_one({"name_key": product_key("Atta")})
    if atta and doc["final_flour_delivered"] > atta.get("current_stock", 0):
        await apply_exchange_effects(old, 1)
        raise HTTPException(status_code=400, detail=f"Not enough Atta stock (have {atta.get('current_stock',0)} kg)")
    await db.exchanges.replace_one({"id": eid}, doc)
    await apply_exchange_effects(doc, 1)
    # Re-settle from scratch: the charge or the way it was paid may have changed.
    for row in await db.payments.find({"ref_id": eid}).to_list(500):
        await unpost_bank_txn(row["id"])
    await db.payments.delete_many({"ref_id": eid})
    await settle_exchange(doc, body)
    await log_audit(user, "Edited exchange", f'{body.customer_name}: {body.wheat_qty} kg wheat, grinding Rs {doc["grinding_charge"]}')
    return clean(doc)

@api_router.delete("/exchanges/{eid}")
async def delete_exchange(eid: str, user: dict = Depends(require_admin)):
    e = await db.exchanges.find_one({"id": eid})
    if e:
        await apply_exchange_effects(e, -1)
        await unpost_material_ledger(eid)
        await db.invoices.delete_one({"ref_id": eid})
        for row in await db.payments.find({"ref_id": eid}).to_list(500):
            await unpost_bank_txn(row["id"])
        await db.payments.delete_many({"ref_id": eid})
        await db.exchanges.delete_one({"id": eid})
    return {"message": "deleted"}

# ---- Record a payment ----
class PayBody(BaseModel):
    payment_method: str = "Cash"
    # How the money moved, and into which account if it was a bank mode.
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None
    reference: str = ""
    # Omitted means "settle the whole balance", preserving the original
    # mark-as-paid behaviour for callers that do not send an amount.
    amount: Optional[float] = None
    date: Optional[str] = None

async def take_payment(coll, rid, method, amount=None, date=None, party="customer", party_field="customer_name", mode="Cash", bank_id=None):
    doc = await db[coll].find_one({"id": rid})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    total = doc.get(TOTAL_FIELD.get(coll, "total"), 0) or 0
    already = await paid_against(rid)
    balance = round(total - already, 2)
    if balance <= 0:
        raise HTTPException(status_code=400, detail="This bill is already settled")
    amt = balance if amount is None else round(float(amount), 2)
    if amt <= 0:
        raise HTTPException(status_code=400, detail="Enter an amount greater than zero")
    if amt > balance + 0.009:
        raise HTTPException(status_code=400, detail=f"Balance is only Rs {balance:.2f}")
    await add_credit(party, doc.get(party_field), amt, date or doc.get("date"), rid,
                     f"Payment {doc.get('invoice_number', '')}".strip(), mode=method)
    await db[coll].update_one({"id": rid}, {"$set": {"payment_method": method}})
    return await sync_payment_state(coll, rid)

async def mark_paid(coll, rid, method):
    await take_payment(coll, rid, method)

@api_router.patch("/sales/{rid}/pay")
async def pay_sale(rid: str, body: PayBody, user: dict = Depends(get_current_user)):
    state = await take_payment("sales", rid, body.payment_method, body.amount, body.date, mode=body.payment_mode, bank_id=body.bank_id)
    await log_audit(user, "Recorded payment", f'sales {rid}: Rs {body.amount if body.amount is not None else "full balance"}')
    return state

@api_router.patch("/grinding/{rid}/pay")
async def pay_grinding(rid: str, body: PayBody, user: dict = Depends(get_current_user)):
    state = await take_payment("grinding", rid, body.payment_method, body.amount, body.date, mode=body.payment_mode, bank_id=body.bank_id)
    await log_audit(user, "Recorded payment", f'grinding {rid}: Rs {body.amount if body.amount is not None else "full balance"}')
    return state

@api_router.patch("/oil/{rid}/pay")
async def pay_oil(rid: str, body: PayBody, user: dict = Depends(get_current_user)):
    state = await take_payment("oil", rid, body.payment_method, body.amount, body.date, mode=body.payment_mode, bank_id=body.bank_id)
    await log_audit(user, "Recorded payment", f'oil {rid}: Rs {body.amount if body.amount is not None else "full balance"}')
    return state

@api_router.patch("/purchases/{rid}/pay")
async def pay_purchase(rid: str, body: PayBody, user: dict = Depends(get_current_user)):
    state = await take_payment("purchases", rid, body.payment_method, body.amount, body.date,
                               party="supplier", party_field="supplier_name",
                               mode=body.payment_mode, bank_id=body.bank_id)
    await log_audit(user, "Recorded payment", f'purchase {rid}: Rs {body.amount if body.amount is not None else "full balance"}')
    return state

# ---- Edit records ----
@api_router.put("/sales/{sid}")
async def edit_sale(sid: str, body: SaleBody, user: dict = Depends(get_current_user)):
    old = await db.sales.find_one({"id": sid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    await db.products.update_one({"id": old["product_id"]}, {"$inc": {"current_stock": old["quantity"]}})
    total = round(body.quantity * body.price, 2)
    await db.sales.update_one({"id": sid}, {"$set": {**body.model_dump(), "total": total}})
    await db.products.update_one({"id": body.product_id}, {"$inc": {"current_stock": -body.quantity}})
    await db.invoices.update_one({"ref_id": sid}, {"$set": {"customer_name": body.customer_name,
        "date": body.date, "total": total}})
    await retag_bill_payments(sid, body.payment_mode, body.bank_id)
    await sync_payment_state("sales", sid)
    return clean(await db.sales.find_one({"id": sid}))

@api_router.put("/grinding/{gid}")
async def edit_grinding(gid: str, body: GrindingBody, user: dict = Depends(get_current_user)):
    old = await db.grinding.find_one({"id": gid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    await apply_grinding_effects(old, -1)
    doc = await build_grinding_doc(body)
    doc["id"] = gid
    doc["invoice_number"] = old["invoice_number"]
    await db.grinding.replace_one({"id": gid}, doc)
    await apply_grinding_effects(doc, 1)
    await db.invoices.update_one({"ref_id": gid}, {"$set": {"customer_name": doc["customer_name"],
        "date": doc["date"], "total": doc["total_charge"]}})
    await retag_bill_payments(gid, body.payment_mode, body.bank_id)
    await sync_payment_state("grinding", gid)
    return clean(await db.grinding.find_one({"id": gid}))

@api_router.put("/oil/{oid}")
async def edit_oil(oid: str, body: OilBody, user: dict = Depends(get_current_user)):
    old = await db.oil.find_one({"id": oid})
    if not old:
        raise HTTPException(status_code=404, detail="Not found")
    await apply_oil_effects(old, -1)
    doc = await build_oil_doc(body)
    doc["id"] = oid
    doc["invoice_number"] = old["invoice_number"]
    await db.oil.replace_one({"id": oid}, doc)
    await apply_oil_effects(doc, 1)
    await db.invoices.update_one({"ref_id": oid}, {"$set": {"customer_name": doc["customer_name"],
        "date": doc["date"], "total": doc["total"]}})
    await retag_bill_payments(oid, body.payment_mode, body.bank_id)
    await sync_payment_state("oil", oid)
    return clean(await db.oil.find_one({"id": oid}))

@api_router.put("/expenses/{eid}")
async def edit_expense(eid: str, body: ExpenseBody, user: dict = Depends(get_current_user)):
    await db.expenses.update_one({"id": eid}, {"$set": body.model_dump()})
    return clean(await db.expenses.find_one({"id": eid}))

# ---- Daybook (end-of-day summary) ----
@api_router.get("/daybook")
async def daybook(date: str, user: dict = Depends(get_current_user)):
    def f(items):
        return [i for i in items if str(i.get("date", "")) == date]
    sales = f(await db.sales.find().to_list(5000))
    grinding = f(await db.grinding.find().to_list(5000))
    oil = f(await db.oil.find().to_list(5000))
    expenses = f(await db.expenses.find().to_list(5000))
    purchases = f(await db.purchases.find().to_list(5000))
    sales_total = sum(s.get("total", 0) for s in sales)
    grinding_total = sum(g.get("total_charge", 0) for g in grinding)
    oil_total = sum(o.get("total", 0) for o in oil)
    income = sales_total + grinding_total + oil_total
    def taken(d, field):
        if d.get("amount_paid") is not None:
            return d.get("amount_paid", 0)
        return d.get(field, 0) if d.get("payment_status") == "Paid" else 0

    collected = (sum(taken(s, "total") for s in sales)
                 + sum(taken(g, "total_charge") for g in grinding)
                 + sum(taken(o, "total") for o in oil))
    exp_total = sum(e.get("amount", 0) for e in expenses)
    return {"date": date, "sales_total": round(sales_total, 2), "grinding_total": round(grinding_total, 2),
            "oil_total": round(oil_total, 2), "income": round(income, 2), "collected": round(collected, 2),
            "pending": round(income - collected, 2), "expenses": round(exp_total, 2),
            "purchases": round(sum(p.get("total", 0) for p in purchases), 2),
            "net": round(collected - exp_total, 2),
            "counts": {"sales": len(sales), "grinding": len(grinding), "oil": len(oil), "expenses": len(expenses)}}

# ==================== Maintenance & Costing ====================

def compute_next_due(last_date, interval_days):
    try:
        d = datetime.fromisoformat(last_date)
    except Exception:
        d = datetime.now(timezone.utc)
    return (d + timedelta(days=int(interval_days))).strftime("%Y-%m-%d")

class MaintenanceBody(BaseModel):
    machine: str
    task: str
    last_service_date: str
    interval_days: int
    notes: str = ""

@api_router.get("/maintenance")
async def get_maintenance(user: dict = Depends(get_current_user)):
    return [clean(m) for m in await db.maintenance.find().sort("next_due_date", 1).to_list(1000)]

@api_router.post("/maintenance")
async def create_maintenance(body: MaintenanceBody, user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(),
           "next_due_date": compute_next_due(body.last_service_date, body.interval_days), "created_at": now_iso()}
    await db.maintenance.insert_one(doc)
    return clean(doc)

@api_router.put("/maintenance/{mid}")
async def update_maintenance(mid: str, body: MaintenanceBody, user: dict = Depends(get_current_user)):
    await db.maintenance.update_one({"id": mid}, {"$set": {**body.model_dump(),
        "next_due_date": compute_next_due(body.last_service_date, body.interval_days)}})
    return clean(await db.maintenance.find_one({"id": mid}))

@api_router.patch("/maintenance/{mid}/serviced")
async def mark_serviced(mid: str, user: dict = Depends(get_current_user)):
    m = await db.maintenance.find_one({"id": mid})
    if not m:
        raise HTTPException(status_code=404, detail="Not found")
    today_str = datetime.now().strftime("%Y-%m-%d")
    await db.maintenance.update_one({"id": mid}, {"$set": {"last_service_date": today_str,
        "next_due_date": compute_next_due(today_str, m["interval_days"])}})
    return clean(await db.maintenance.find_one({"id": mid}))

@api_router.delete("/maintenance/{mid}")
async def delete_maintenance(mid: str, user: dict = Depends(require_admin)):
    await db.maintenance.delete_one({"id": mid})
    return {"message": "deleted"}

@api_router.get("/costing")
async def costing(user: dict = Depends(get_current_user)):
    finished = {"Flour", "Bran", "Edible Oil", "Oil Cake", "Masala"}
    rows = []
    for p in await db.products.find().sort("name", 1).to_list(1000):
        if p.get("category") in finished:
            cost = round(p.get("cost_per_unit", 0), 2)
            rate = round(p.get("rate", 0), 2)
            margin = round(rate - cost, 2)
            pct = round((margin / rate * 100), 1) if rate else 0
            rows.append({"id": p["id"], "name": p["name"], "category": p["category"], "unit": p.get("unit", "kg"),
                         "current_stock": p.get("current_stock", 0), "cost_per_unit": cost,
                         "rate": rate, "margin": margin, "margin_pct": pct})
    return rows

# ==================== Ledger, Cash Book, Analytics, Audit ====================

# How the money moved, and therefore which ledger it lands in. Only Cash touches
# the drawer; every bank mode settles into an account, so the cash book has to
# tell them apart or Cash in Hand drifts up by every digital payment ever taken.
BANK_MODES = ("Bank", "UPI", "NEFT", "RTGS", "IMPS", "Cheque")

# Paid in kind — flour kept back as the grinding fee, or grain handed over
# instead of cash. Moves stock, never money, so it must reach neither ledger.
MATERIAL_MODES = ("Grain", "Material", "Flour")

PAYMENT_MODES = ("Cash",) + BANK_MODES + MATERIAL_MODES

# Written by the brief period when only Cash and Bank were offered, and by the
# older free-text fields. Mapped rather than dropped because clean_mode runs on
# read as well as write: an unrecognised value falls through to Cash, which
# would silently reclassify past digital payments as drawer cash.
MODE_ALIASES = {
    "BANK TRANSFER": "Bank", "NET BANKING": "Bank", "NETBANKING": "Bank",
    "CHECK": "Cheque", "CHQ": "Cheque", "DD": "Cheque",
    "G-PAY": "UPI", "GPAY": "UPI", "PHONEPE": "UPI", "PAYTM": "UPI",
}

_MODE_LOOKUP = {m.upper(): m for m in PAYMENT_MODES}

def clean_mode(mode) -> str:
    raw = (mode or "Cash").strip()
    key = raw.upper()
    return _MODE_LOOKUP.get(key) or MODE_ALIASES.get(key) or "Cash"

def is_bank_mode(mode) -> bool:
    return clean_mode(mode) in BANK_MODES

def is_material_mode(mode) -> bool:
    return clean_mode(mode) in MATERIAL_MODES

def party_key(name: str) -> str:
    return " ".join((name or "").split()).lower()

async def find_party(coll: str, name: str):
    return await db[coll].find_one({"party_key": party_key(name)})

async def ensure_party(coll: str, name: str) -> dict:
    """Return the named party, creating a bare record if it is new.

    Transactions can name a party that was never entered in the master — a
    walk-in, or a typo corrected later. Creating the record keeps every ledger
    attached to exactly one party instead of stranding history under a name
    with no master entry.
    """
    key = party_key(name)
    if not key:
        return {}
    found = await db[coll].find_one({"party_key": key})
    if found:
        return clean(found)
    doc = {"id": str(uuid.uuid4()), "name": name.strip(), "party_key": key,
           "phone": "", "address": "", "gstin": "", "pan_aadhaar": "",
           "opening_balance": 0, "credit_limit": 0, "created_at": now_iso()}
    await db[coll].insert_one(doc)
    return clean(doc)

async def add_credit(party_type, name, amount, date, ref_id=None, note="", kind="receipt", mode="Cash", bank_id=None):
    """Record money moving against a party.

    A negative amount on a customer is money paid out to them, which keeps the
    ledger arithmetic honest: outstanding is debits minus credits, so a -138
    credit against a -138 bill settles to zero. `kind` marks it as a refund so
    the cash book can report it as an outflow rather than negative income.

    `mode` is stored per payment, not per bill: a customer can settle half in
    cash today and half by bank transfer next week, and the cash book needs both.
    """
    if not amount:
        return
    resolved = clean_mode(mode)
    pid = str(uuid.uuid4())
    await db.payments.insert_one({"id": pid, "party_type": party_type, "party_name": name,
        "amount": round(amount, 2), "date": date, "note": note, "ref_id": ref_id,
        "kind": kind, "payment_mode": resolved, "bank_id": bank_id, "created_at": now_iso()})
    await post_payment_to_bank(pid)
    return pid

async def retag_bill_payments(ref_id: str, mode, bank_id=None):
    """Move a bill's own payments onto the mode it now says it was settled by.

    Editing a bill from NEFT to Cash means the money did not come through the
    bank after all. Without this the payment keeps its old mode and its bank row
    survives, leaving the account permanently overstated.
    """
    resolved = clean_mode(mode)
    rows = await db.payments.find({"ref_id": ref_id}).to_list(500)
    for row in rows:
        await db.payments.update_one({"id": row["id"]},
                                     {"$set": {"payment_mode": resolved, "bank_id": bank_id}})
        await post_payment_to_bank(row["id"])

async def post_payment_to_bank(payment_id: str):
    """Mirror a bank-mode payment into the bank ledger, exactly once.

    This is what makes the posting automatic: nothing else in the app has to
    remember to write a bank row. Keyed on the payment id, so re-running after
    an edit updates the existing row rather than adding a second, and a payment
    that changes from UPI to Cash has its bank row removed instead of stranded.
    """
    pay = await db.payments.find_one({"id": payment_id})
    if not pay:
        return
    mode = clean_mode(pay.get("payment_mode"))
    # Cash never reaches an account, and paying in flour or grain moves stock
    # rather than money. Neither belongs in a bank ledger.
    if mode not in BANK_MODES:
        await unpost_bank_txn(payment_id)
        return
    bank_id = pay.get("bank_id") or await default_bank_id()
    if not bank_id:
        # Several accounts and none chosen: recording it against a guess would
        # put money in the wrong ledger. The payment stands; the posting waits.
        logger.warning("Payment %s is %s but no bank account was selected; not posted.", payment_id, mode)
        return
    amount = pay.get("amount", 0) or 0
    party = pay.get("party_type")
    # A positive customer row is money received; a negative one is a refund paid
    # out. Supplier rows are always money leaving.
    inflow = party == "customer" and amount >= 0
    await post_bank_txn(
        bank_id=bank_id, date=pay.get("date"),
        txn_type="Bank Receipt" if inflow else "Bank Payment",
        amount=abs(amount), mode=mode, party_name=pay.get("party_name", ""),
        reference=pay.get("reference", ""), note=pay.get("note", ""),
        source_ref=payment_id, source_kind="payment")

# ---------------- Part payment ----------------
# A bill can be settled over several visits. The payments collection is the
# single source of truth for how much has come in; amount_paid on the record is
# a cached sum of it, recomputed rather than incremented so an edit or a
# repeated call can never double-count.

TOTAL_FIELD = {"sales": "total", "grinding": "total_charge", "oil": "total", "purchases": "total"}

def payment_state(total: float, paid: float) -> str:
    # A negative total means the by-product the customer sold us is worth more
    # than the service charge, so the shop hands cash over instead of taking it.
    # That is not a debt owed to us and must never read Pending.
    if (total or 0) < 0:
        return "Paid to Customer"
    if paid <= 0:
        return "Pending"
    if paid + 0.009 >= (total or 0):
        return "Paid"
    return "Partial"

async def paid_against(ref_id: str) -> float:
    rows = await db.payments.find({"ref_id": ref_id}).to_list(500)
    return round(sum(r.get("amount", 0) for r in rows), 2)

async def sync_payment_state(coll: str, rid: str) -> dict:
    """Recompute amount_paid/balance/status for a record from its payments."""
    doc = await db[coll].find_one({"id": rid})
    if not doc:
        return {}
    total = doc.get(TOTAL_FIELD.get(coll, "total"), 0) or 0
    paid = await paid_against(rid)
    status = payment_state(total, paid)
    # paid_to_customer carries the amount owed the other way, so the UI has a
    # positive figure to show and nothing has to infer it from a negative total.
    state = {
        "amount_paid": paid,
        "balance_due": round(max(total - paid, 0), 2) if total >= 0 else 0.0,
        "paid_to_customer": round(abs(total), 2) if total < 0 else 0.0,
        "payment_status": status,
    }
    await db[coll].update_one({"id": rid}, {"$set": state})
    await db.invoices.update_one({"ref_id": rid}, {"$set": state})
    return state

async def log_audit(user, action, detail=""):
    u = user or {}
    await db.audit.insert_one({"id": str(uuid.uuid4()), "user": u.get("name") or u.get("email") or "system",
        "role": u.get("role", ""), "action": action, "detail": detail, "at": now_iso()})

@api_router.get("/audit")
async def get_audit(user: dict = Depends(get_current_user)):
    return [clean(a) for a in await db.audit.find().sort("at", -1).to_list(500)]

class PaymentBody(BaseModel):
    party_type: str
    party_name: str
    amount: float
    date: str
    note: str = ""
    payment_mode: str = "Cash"
    bank_id: Optional[str] = None
    reference: str = ""

@api_router.get("/payments")
async def list_payments(party_type: Optional[str] = None, party_name: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if party_type:
        q["party_type"] = party_type
    if party_name:
        q["party_name"] = party_name
    return [clean(p) for p in await db.payments.find(q).sort("date", -1).to_list(5000)]

@api_router.post("/payments")
async def create_payment(body: PaymentBody, user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "ref_id": None, "created_at": now_iso()}
    await db.payments.insert_one(doc)
    await log_audit(user, "Recorded payment", f"{body.party_type} {body.party_name}: Rs {body.amount}")
    return clean(doc)

@api_router.delete("/payments/{pid}")
async def delete_payment(pid: str, user: dict = Depends(require_admin)):
    await db.payments.delete_one({"id": pid})
    return {"message": "deleted"}

async def build_ledger(party_type, name):
    entries = []
    if party_type == "customer":
        for coll, field, label in [("sales", "total", "Sale"), ("grinding", "total_charge", "Grinding"), ("oil", "total", "Oil Extraction")]:
            for d in await db[coll].find({"customer_name": name}).to_list(3000):
                if d.get(field, 0):
                    entries.append({"date": d.get("date"), "type": label, "ref": d.get("invoice_number", ""), "debit": round(d.get(field, 0), 2), "credit": 0})
    else:
        for d in await db.purchases.find({"supplier_name": name}).to_list(3000):
            entries.append({"date": d.get("date"), "type": "Purchase", "ref": d.get("product_name", ""), "debit": round(d.get("total", 0), 2), "credit": 0})
    for p in await db.payments.find({"party_type": party_type, "party_name": name}).to_list(3000):
        entries.append({"date": p.get("date"), "type": "Payment", "ref": p.get("note", ""), "debit": 0, "credit": round(p.get("amount", 0), 2)})
    entries.sort(key=lambda e: (e.get("date") or ""))
    bal = 0.0
    for e in entries:
        bal += e["debit"] - e["credit"]
        e["balance"] = round(bal, 2)
    return {"name": name, "entries": entries, "total_debit": round(sum(e["debit"] for e in entries), 2),
            "total_credit": round(sum(e["credit"] for e in entries), 2), "balance": round(bal, 2)}

@api_router.get("/customers/{cid}/ledger")
async def customer_ledger(cid: str, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one({"id": cid})
    if not c:
        raise HTTPException(status_code=404, detail="Not found")
    return await build_ledger("customer", c["name"])

@api_router.get("/suppliers/{sid}/ledger")
async def supplier_ledger(sid: str, user: dict = Depends(get_current_user)):
    s = await db.suppliers.find_one({"id": sid})
    if not s:
        raise HTTPException(status_code=404, detail="Not found")
    return await build_ledger("supplier", s["name"])

@api_router.get("/cashbook")
async def cashbook(date: str, user: dict = Depends(get_current_user)):
    settings = await get_settings_doc()
    starting = settings.get("starting_cash", 0) or 0
    payments = await db.payments.find().to_list(20000)
    expenses = await db.expenses.find().to_list(20000)
    # Customer rows with a negative amount are cash handed back, so split them
    # out rather than letting them net against receipts and read as less income.
    cust_rows = [p for p in payments if p.get("party_type") == "customer"]
    cust_in = [p for p in cust_rows if p.get("amount", 0) >= 0]
    cust_refund = [p for p in cust_rows if p.get("amount", 0) < 0]
    supp_out = [p for p in payments if p.get("party_type") == "supplier"]

    # Only notes and coins move the drawer. A bank transfer is real
    # income but never reaches it, so counting it as cash would inflate Cash in
    # Hand by every digital payment ever taken. Rows written before payment_mode
    # existed are treated as cash, which is what they were assumed to be.
    def is_cash(row):
        return clean_mode(row.get("payment_mode")) == "Cash"

    def cash(items):
        return [i for i in items if is_cash(i)]

    def digital(items):
        # Bank modes only. Grain and flour settle in kind and never touch a
        # money ledger, so they must not appear as bank movement either.
        return [i for i in items if is_bank_mode(i.get("payment_mode"))]

    def before(items):
        return [i for i in items if str(i.get("date", "")) < date]

    def on(items):
        return [i for i in items if str(i.get("date", "")) == date]

    def total(items):
        return sum(abs(i.get("amount", 0)) for i in items)

    opening = (starting
               + total(cash(before(cust_in)))
               - total(cash(before(cust_refund)))
               - total(cash(before(supp_out)))
               - sum(e.get("amount", 0) for e in before(expenses)))

    in_today = total(cash(on(cust_in)))
    refund_today = total(cash(on(cust_refund)))
    supp_today = total(cash(on(supp_out)))
    exp_today = sum(e.get("amount", 0) for e in on(expenses))
    closing = opening + in_today - refund_today - supp_today - exp_today

    # Digital movements are reported alongside so the day still reconciles
    # against what actually came in, it just does not touch the drawer.
    bank_in_today = total(digital(on(cust_in)))
    bank_out_today = total(digital(on(cust_refund))) + total(digital(on(supp_out)))

    by_mode = {}
    for row in on(cust_in):
        by_mode[clean_mode(row.get("payment_mode"))] = round(
            by_mode.get(clean_mode(row.get("payment_mode")), 0) + abs(row.get("amount", 0)), 2)

    return {"date": date, "opening": round(opening, 2), "payments_received": round(in_today, 2),
            "paid_to_customers": round(refund_today, 2),
            "supplier_payments": round(supp_today, 2), "expenses": round(exp_today, 2),
            "closing": round(closing, 2),
            "bank_received": round(bank_in_today, 2), "bank_paid": round(bank_out_today, 2),
            "received_by_mode": by_mode,
            "total_received": round(in_today + bank_in_today, 2)}

# ==================== Bank Management ====================
# Balances are always derived: opening_balance plus the sum of the account's
# transactions. A stored running balance would drift the moment any transaction
# was edited, deleted or posted out of order, and this ledger has to reconcile
# against a real bank statement.

ACCOUNT_TYPES = ("Savings", "Current", "Cash Credit")

# Signed so the arithmetic is uniform: inflows add, outflows subtract, and a
# balance is just a sum. Transfers are recorded as a pair.
TXN_SIGN = {
    "Deposit": 1, "Bank Receipt": 1, "Interest Credit": 1, "Transfer In": 1,
    "Withdrawal": -1, "Bank Payment": -1, "Bank Charges": -1, "Transfer Out": -1,
}

class BankAccountBody(BaseModel):
    bank_name: str
    branch: str = ""
    account_number: str = ""
    ifsc: str = ""
    holder_name: str = ""
    opening_balance: float = 0
    account_type: str = "Current"

class BankTxnBody(BaseModel):
    bank_id: str
    date: str
    txn_type: str
    amount: float
    mode: str = "Bank"
    party_name: str = ""
    reference: str = ""          # cheque number, UTR, UPI reference
    note: str = ""

class BankTransferBody(BaseModel):
    from_bank_id: str
    to_bank_id: str
    date: str
    amount: float
    reference: str = ""
    note: str = ""

class ReconcileBody(BaseModel):
    reconciled: bool = True
    reconciled_date: Optional[str] = None

def account_digits(acc: str) -> str:
    return "".join(ch for ch in (acc or "") if ch.isdigit())

async def bank_balance(bank_id: str) -> float:
    acc = await db.bank_accounts.find_one({"id": bank_id})
    if not acc:
        return 0.0
    rows = await db.bank_txns.find({"bank_id": bank_id}).to_list(50000)
    return round((acc.get("opening_balance", 0) or 0) + sum(r.get("amount", 0) for r in rows), 2)

async def default_bank_id() -> Optional[str]:
    """The account to post to when a payment names a bank mode but no account.

    With exactly one account there is no ambiguity, so posting is automatic as
    the spec requires. With several, guessing would put money in the wrong
    ledger, so the caller must choose; the payment is still recorded and the
    posting is left for the operator to place.
    """
    accounts = await db.bank_accounts.find({"active": {"$ne": False}}).to_list(100)
    if len(accounts) == 1:
        return accounts[0]["id"]
    marked = next((a for a in accounts if a.get("is_default")), None)
    return marked["id"] if marked else None

async def post_bank_txn(*, bank_id, date, txn_type, amount, mode="Bank", party_name="",
                        reference="", note="", source_ref=None, source_kind=None):
    """Write one bank transaction, at most once per source.

    source_ref carries the id of whatever caused this posting — a payment, a
    grinding bill. The unique index on it is what makes automatic posting safe
    to re-run: an edit or a retry updates the existing row instead of adding a
    second one, which is the "no duplicate ledger entries" requirement.
    """
    signed = round(abs(amount) * TXN_SIGN.get(txn_type, 1), 2)
    doc = {"bank_id": bank_id, "date": date, "txn_type": txn_type, "amount": signed,
           "mode": clean_mode(mode), "party_name": party_name, "reference": reference,
           "note": note, "source_kind": source_kind,
           "reconciled": False, "reconciled_date": None}
    # Only set when there is a source. A stored null still counts as a value for
    # the unique index, so every manual row would collide on null with the first.
    if source_ref:
        doc["source_ref"] = source_ref
    if source_ref:
        existing = await db.bank_txns.find_one({"source_ref": source_ref})
        if existing:
            # Keep the reconciliation marks; only the money details change.
            doc["reconciled"] = existing.get("reconciled", False)
            doc["reconciled_date"] = existing.get("reconciled_date")
            await db.bank_txns.update_one({"id": existing["id"]}, {"$set": doc})
            return existing["id"]
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    await db.bank_txns.insert_one(doc)
    return doc["id"]

async def unpost_bank_txn(source_ref: str):
    """Remove an automatic posting when its source is deleted or turns cash."""
    if source_ref:
        await db.bank_txns.delete_many({"source_ref": source_ref})

@api_router.get("/banks")
async def list_banks(user: dict = Depends(get_current_user)):
    accounts = [clean(a) for a in await db.bank_accounts.find().sort("bank_name", 1).to_list(200)]
    for a in accounts:
        a["balance"] = await bank_balance(a["id"])
    return accounts

@api_router.post("/banks")
async def create_bank(body: BankAccountBody, user: dict = Depends(require_admin)):
    if not body.bank_name.strip():
        raise HTTPException(status_code=400, detail="Enter the bank name")
    acc_type = body.account_type if body.account_type in ACCOUNT_TYPES else "Current"
    digits = account_digits(body.account_number)
    if digits and await db.bank_accounts.find_one({"account_digits": digits}):
        raise HTTPException(status_code=400, detail="An account with that number already exists")
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "account_type": acc_type,
           "account_digits": digits, "active": True, "created_at": now_iso()}
    await db.bank_accounts.insert_one(doc)
    await log_audit(user, "Added bank account", f'{body.bank_name} {body.account_number}')
    out = clean(doc)
    out["balance"] = round(body.opening_balance or 0, 2)
    return out

@api_router.put("/banks/{bid}")
async def update_bank(bid: str, body: BankAccountBody, user: dict = Depends(require_admin)):
    acc = await db.bank_accounts.find_one({"id": bid})
    if not acc:
        raise HTTPException(status_code=404, detail="Bank account not found")
    acc_type = body.account_type if body.account_type in ACCOUNT_TYPES else "Current"
    digits = account_digits(body.account_number)
    clash = await db.bank_accounts.find_one({"account_digits": digits, "id": {"$ne": bid}}) if digits else None
    if clash:
        raise HTTPException(status_code=400, detail="An account with that number already exists")
    await db.bank_accounts.update_one({"id": bid}, {"$set": {**body.model_dump(),
        "account_type": acc_type, "account_digits": digits}})
    await log_audit(user, "Edited bank account", body.bank_name)
    out = clean(await db.bank_accounts.find_one({"id": bid}))
    out["balance"] = await bank_balance(bid)
    return out

@api_router.delete("/banks/{bid}")
async def delete_bank(bid: str, user: dict = Depends(require_admin)):
    # Transactions are the audit trail for money that really moved, so an
    # account carrying any is closed rather than erased.
    if await db.bank_txns.count_documents({"bank_id": bid}):
        await db.bank_accounts.update_one({"id": bid}, {"$set": {"active": False}})
        await log_audit(user, "Closed bank account", bid)
        return {"message": "closed", "detail": "Account has transactions, so it was closed rather than deleted."}
    await db.bank_accounts.delete_one({"id": bid})
    await log_audit(user, "Deleted bank account", bid)
    return {"message": "deleted"}

@api_router.get("/bank-transactions")
async def list_bank_txns(bank_id: Optional[str] = None, start: Optional[str] = None,
                         end: Optional[str] = None, mode: Optional[str] = None,
                         txn_type: Optional[str] = None, reconciled: Optional[bool] = None,
                         user: dict = Depends(get_current_user)):
    q = {}
    if bank_id: q["bank_id"] = bank_id
    if mode: q["mode"] = clean_mode(mode)
    if txn_type: q["txn_type"] = txn_type
    if reconciled is not None: q["reconciled"] = reconciled
    if start or end:
        rng = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end
        q["date"] = rng
    rows = [clean(r) for r in await db.bank_txns.find(q).sort("date", -1).to_list(20000)]
    names = {a["id"]: a.get("bank_name", "") for a in await db.bank_accounts.find().to_list(200)}
    for r in rows:
        r["bank_name"] = names.get(r.get("bank_id"), "")
    return rows

@api_router.post("/bank-transactions")
async def create_bank_txn(body: BankTxnBody, user: dict = Depends(get_current_user)):
    if body.txn_type not in TXN_SIGN:
        raise HTTPException(status_code=400, detail=f"Unknown transaction type: {body.txn_type}")
    if not await db.bank_accounts.find_one({"id": body.bank_id}):
        raise HTTPException(status_code=404, detail="Bank account not found")
    if abs(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Enter an amount greater than zero")
    tid = await post_bank_txn(bank_id=body.bank_id, date=body.date, txn_type=body.txn_type,
                              amount=body.amount, mode=body.mode, party_name=body.party_name,
                              reference=body.reference, note=body.note)
    await log_audit(user, "Bank transaction", f'{body.txn_type} Rs {abs(body.amount)}')
    row = clean(await db.bank_txns.find_one({"id": tid}))
    row["balance"] = await bank_balance(body.bank_id)
    return row

@api_router.delete("/bank-transactions/{tid}")
async def delete_bank_txn(tid: str, user: dict = Depends(require_admin)):
    row = await db.bank_txns.find_one({"id": tid})
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    if row.get("source_ref"):
        raise HTTPException(status_code=400,
            detail="This was posted automatically from a payment. Edit or delete that record instead.")
    await db.bank_txns.delete_one({"id": tid})
    await log_audit(user, "Deleted bank transaction", f'{row.get("txn_type")} Rs {abs(row.get("amount", 0))}')
    return {"message": "deleted"}

@api_router.post("/bank-transactions/transfer")
async def bank_transfer(body: BankTransferBody, user: dict = Depends(get_current_user)):
    if body.from_bank_id == body.to_bank_id:
        raise HTTPException(status_code=400, detail="Choose two different accounts")
    if abs(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Enter an amount greater than zero")
    src = await db.bank_accounts.find_one({"id": body.from_bank_id})
    dst = await db.bank_accounts.find_one({"id": body.to_bank_id})
    if not src or not dst:
        raise HTTPException(status_code=404, detail="Bank account not found")
    # Paired rows sharing a transfer_id, so a report can show one movement while
    # each account's own balance stays a simple sum of its rows.
    transfer_id = str(uuid.uuid4())
    out_id = await post_bank_txn(bank_id=body.from_bank_id, date=body.date, txn_type="Transfer Out",
                                 amount=body.amount, mode="Bank", party_name=dst.get("bank_name", ""),
                                 reference=body.reference, note=body.note or f'Transfer to {dst.get("bank_name","")}')
    in_id = await post_bank_txn(bank_id=body.to_bank_id, date=body.date, txn_type="Transfer In",
                                amount=body.amount, mode="Bank", party_name=src.get("bank_name", ""),
                                reference=body.reference, note=body.note or f'Transfer from {src.get("bank_name","")}')
    await db.bank_txns.update_many({"id": {"$in": [out_id, in_id]}}, {"$set": {"transfer_id": transfer_id}})
    await log_audit(user, "Bank transfer", f'{src.get("bank_name")} to {dst.get("bank_name")} Rs {abs(body.amount)}')
    return {"transfer_id": transfer_id,
            "from_balance": await bank_balance(body.from_bank_id),
            "to_balance": await bank_balance(body.to_bank_id)}

@api_router.patch("/bank-transactions/{tid}/reconcile")
async def reconcile_txn(tid: str, body: ReconcileBody, user: dict = Depends(get_current_user)):
    row = await db.bank_txns.find_one({"id": tid})
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    await db.bank_txns.update_one({"id": tid}, {"$set": {
        "reconciled": body.reconciled,
        "reconciled_date": (body.reconciled_date or datetime.now(timezone.utc).strftime("%Y-%m-%d")) if body.reconciled else None}})
    return clean(await db.bank_txns.find_one({"id": tid}))

@api_router.get("/bank-reconciliation")
async def bank_reconciliation(bank_id: str, as_of: Optional[str] = None,
                              statement_balance: Optional[float] = None,
                              user: dict = Depends(get_current_user)):
    """Book balance against the statement, and what accounts for the gap."""
    acc = await db.bank_accounts.find_one({"id": bank_id})
    if not acc:
        raise HTTPException(status_code=404, detail="Bank account not found")
    q = {"bank_id": bank_id}
    if as_of:
        q["date"] = {"$lte": as_of}
    rows = [clean(r) for r in await db.bank_txns.find(q).sort("date", 1).to_list(20000)]
    opening = acc.get("opening_balance", 0) or 0
    book = round(opening + sum(r.get("amount", 0) for r in rows), 2)
    unreconciled = [r for r in rows if not r.get("reconciled")]
    # Cleared balance is what the bank should be showing: the book balance less
    # anything not yet through, which is exactly the reconciling difference.
    uncleared = round(sum(r.get("amount", 0) for r in unreconciled), 2)
    cleared = round(book - uncleared, 2)
    out = {"bank_id": bank_id, "bank_name": acc.get("bank_name"), "as_of": as_of,
           "opening_balance": round(opening, 2), "book_balance": book,
           "cleared_balance": cleared, "uncleared_total": uncleared,
           "reconciled_count": len(rows) - len(unreconciled),
           "unreconciled_count": len(unreconciled), "unreconciled": unreconciled}
    if statement_balance is not None:
        out["statement_balance"] = round(statement_balance, 2)
        out["difference"] = round(statement_balance - cleared, 2)
        out["matched"] = abs(out["difference"]) < 0.01
    return out

@api_router.get("/bank-summary")
async def bank_summary(user: dict = Depends(get_current_user)):
    accounts = [clean(a) for a in await db.bank_accounts.find().sort("bank_name", 1).to_list(200)]
    for a in accounts:
        a["balance"] = await bank_balance(a["id"])
        a["unreconciled_count"] = await db.bank_txns.count_documents({"bank_id": a["id"], "reconciled": False})
    return {"accounts": accounts,
            "total_balance": round(sum(a["balance"] for a in accounts if a.get("active") is not False), 2)}

# ==================== Reports ====================
# Every report returns the same shape — columns, rows, summary — so the date
# filters and the print, PDF and Excel exports are written once rather than
# fourteen times. A new report is a builder function and a registry entry.

def date_range(preset: str = None, start: str = None, end: str = None):
    """Resolve a preset or an explicit range into (start, end) inclusive."""
    today = datetime.now(timezone.utc).date()
    p = (preset or "").lower().replace(" ", "-")
    if p == "today":
        return str(today), str(today)
    if p == "yesterday":
        y = today - timedelta(days=1)
        return str(y), str(y)
    if p in ("this-week", "week"):
        return str(today - timedelta(days=today.weekday())), str(today)
    if p in ("this-month", "month"):
        return str(today.replace(day=1)), str(today)
    if p in ("financial-year", "fy"):
        # Indian financial year runs April to March.
        fy_start = today.replace(month=4, day=1) if today.month >= 4 else today.replace(year=today.year - 1, month=4, day=1)
        return str(fy_start), str(today)
    return (start or "1900-01-01"), (end or str(today))

def in_range(row, start, end, field="date"):
    d = str(row.get(field, "") or "")
    return bool(d) and start <= d <= end

def col(key, label, kind="text", align=None):
    return {"key": key, "label": label, "type": kind,
            "align": align or ("right" if kind in ("money", "qty") else "left")}

async def party_movements(name: str):
    """Every debit and credit against one party, oldest first."""
    rows = []
    for coll, field, label in (("sales", "total", "Sale"), ("grinding", "total_charge", "Grinding"),
                               ("oil", "total", "Oil extraction"), ("exchanges", "grinding_charge", "Exchange")):
        for d in await db[coll].find({"customer_name": name}).to_list(5000):
            if d.get(field):
                rows.append({"date": d.get("date"), "particulars": f'{label} {d.get("invoice_number", "")}'.strip(),
                             "debit": round(d.get(field, 0), 2), "credit": 0})
    for d in await db.purchases.find({"supplier_name": name}).to_list(5000):
        rows.append({"date": d.get("date"), "particulars": f'Purchase {d.get("product_name", "")}',
                     "debit": round(d.get("total", 0), 2), "credit": 0})
    for pay in await db.payments.find({"party_name": name}).to_list(5000):
        amt = pay.get("amount", 0) or 0
        note = pay.get("note") or "Payment"
        mode = clean_mode(pay.get("payment_mode"))
        if amt >= 0:
            rows.append({"date": pay.get("date"), "particulars": f'{note} ({mode})', "debit": 0, "credit": round(amt, 2)})
        else:
            # Money paid back out to a customer is a debit on their account.
            rows.append({"date": pay.get("date"), "particulars": f'{note} ({mode})', "debit": round(abs(amt), 2), "credit": 0})
    rows.sort(key=lambda r: str(r.get("date") or ""))
    return rows

async def product_movements(start: str, end: str):
    """Stock in and out per product, split by what caused it.

    Opening is derived by walking today's figure back through everything dated
    on or after the start, because the app stores a running stock rather than a
    movement history. Adjustments are included, which is why they are recorded
    with a date.
    """
    products = [clean(p) for p in await db.products.find().sort("name", 1).to_list(2000)]
    by_name = {p["name"]: p for p in products}
    acc = {p["name"]: {"purchased": 0.0, "sold": 0.0, "produced": 0.0, "consumed": 0.0,
                       "adjusted": 0.0, "after_start": 0.0} for p in products}

    def touch(name, qty, bucket, date):
        if name not in acc:
            return
        if start <= str(date or "") <= end:
            acc[name][bucket] += qty
        if str(date or "") >= start:
            acc[name]["after_start"] += qty

    for d in await db.purchases.find().to_list(20000):
        touch(d.get("product_name"), d.get("quantity", 0) or 0, "purchased", d.get("date"))
    for d in await db.sales.find().to_list(20000):
        touch(d.get("product_name"), -(d.get("quantity", 0) or 0), "sold", d.get("date"))
    for d in await db.production.find().to_list(20000):
        touch(d.get("input_product_name"), -(d.get("input_quantity", 0) or 0), "consumed", d.get("date"))
        for o in d.get("outputs", []):
            touch(o.get("product_name"), o.get("quantity", 0) or 0, "produced", d.get("date"))
    for d in await db.grinding.find().to_list(20000):
        method = normalise_method(d.get("payment_method"))
        if method == FLOUR_DEDUCTION and d.get("deducted_flour"):
            touch(grinding_output_name(d), d["deducted_flour"], "produced", d.get("date"))
        elif method == GRAIN_DEDUCTION and d.get("grain_qty"):
            touch(d.get("grain_item") or "Wheat Crop", d["grain_qty"], "purchased", d.get("date"))
    for d in await db.exchanges.find().to_list(20000):
        touch("Wheat Crop", d.get("wheat_qty", 0) or 0, "purchased", d.get("date"))
        touch("Atta", -(d.get("final_flour_delivered", d.get("atta_given", 0)) or 0), "sold", d.get("date"))
    for d in await db.oil.find().to_list(20000):
        for field, name in (("retained_oil", "Mustard Oil"), ("retained_cake", "Mustard Oil Cake"),
                            ("cake_sold_to_shop", "Mustard Oil Cake")):
            if d.get(field):
                touch(name, d[field], "purchased", d.get("date"))
    for d in await db.stock_adjustments.find().to_list(20000):
        touch(d.get("product_name"), d.get("delta", 0) or 0, "adjusted", d.get("date"))

    rows = []
    for name, a in acc.items():
        prod = by_name[name]
        current = prod.get("current_stock", 0) or 0
        opening = round(current - a["after_start"], 3)
        moved = a["purchased"] + a["sold"] + a["produced"] + a["consumed"] + a["adjusted"]
        rows.append({
            "item": name, "unit": prod.get("unit", "kg"), "category": prod.get("category", ""),
            "opening": opening,
            "purchased": round(a["purchased"], 3),
            "sold": round(abs(a["sold"]), 3),
            "produced": round(a["produced"], 3),
            "consumed": round(abs(a["consumed"]), 3),
            "adjusted": round(a["adjusted"], 3),
            "closing": round(opening + moved, 3),
            "current": round(current, 3),
            "rate": prod.get("rate", 0), "cost_per_unit": prod.get("cost_per_unit", 0),
            "value": round(current * (prod.get("cost_per_unit", 0) or 0), 2),
        })
    return rows

async def rep_sales(start, end, party=None, item=None, **_):
    rows = [clean(d) for d in await db.sales.find().sort("date", -1).to_list(20000)]
    rows = [r for r in rows if in_range(r, start, end)]
    if party: rows = [r for r in rows if r.get("customer_name") == party]
    if item: rows = [r for r in rows if r.get("product_name") == item]
    units = {p["name"]: p.get("unit", "kg") for p in await db.products.find().to_list(2000)}
    out = [{"date": r.get("date"), "invoice": r.get("invoice_number", ""), "customer": r.get("customer_name", ""),
            "item": r.get("product_name", ""), "qty": f'{r.get("quantity", 0)} {units.get(r.get("product_name"), "")}'.strip(),
            "rate": r.get("price", 0), "total": r.get("total", 0),
            "paid": r.get("amount_paid", 0) or 0, "balance": r.get("balance_due", 0) or 0,
            "status": r.get("payment_status", "")} for r in rows]
    return {"columns": [col("date", "Date"), col("invoice", "Invoice"), col("customer", "Customer"),
                        col("item", "Item"), col("qty", "Qty", "qty"), col("rate", "Rate", "money"),
                        col("total", "Total", "money"), col("paid", "Paid", "money"),
                        col("balance", "Balance", "money"), col("status", "Status")],
            "rows": out,
            "summary": [{"label": "Sales", "value": round(sum(r["total"] for r in out), 2), "type": "money"},
                        {"label": "Received", "value": round(sum(r["paid"] for r in out), 2), "type": "money"},
                        {"label": "Outstanding", "value": round(sum(r["balance"] for r in out), 2), "type": "money"},
                        {"label": "Bills", "value": len(out)}]}

async def rep_purchases(start, end, party=None, item=None, **_):
    rows = [clean(d) for d in await db.purchases.find().sort("date", -1).to_list(20000)]
    rows = [r for r in rows if in_range(r, start, end)]
    if party: rows = [r for r in rows if r.get("supplier_name") == party]
    if item: rows = [r for r in rows if r.get("product_name") == item]
    out = [{"date": r.get("date"), "supplier": r.get("supplier_name", ""), "item": r.get("product_name", ""),
            "qty": r.get("quantity", 0), "rate": r.get("rate", 0), "total": r.get("total", 0),
            "paid": r.get("amount_paid", 0) or 0, "balance": r.get("balance_due", 0) or 0,
            "status": r.get("payment_status", "")} for r in rows]
    return {"columns": [col("date", "Date"), col("supplier", "Supplier"), col("item", "Item"),
                        col("qty", "Qty", "qty"), col("rate", "Rate", "money"), col("total", "Total", "money"),
                        col("paid", "Paid", "money"), col("balance", "Balance", "money"), col("status", "Status")],
            "rows": out,
            "summary": [{"label": "Purchases", "value": round(sum(r["total"] for r in out), 2), "type": "money"},
                        {"label": "Paid", "value": round(sum(r["paid"] for r in out), 2), "type": "money"},
                        {"label": "Due to suppliers", "value": round(sum(r["balance"] for r in out), 2), "type": "money"}]}

async def rep_grinding(start, end, party=None, **_):
    rows = [clean(d) for d in await db.grinding.find().sort("date", -1).to_list(20000)]
    rows = [r for r in rows if in_range(r, start, end)]
    if party: rows = [r for r in rows if r.get("customer_name") == party]
    out = [{"date": r.get("date"), "invoice": r.get("invoice_number", ""), "customer": r.get("customer_name", ""),
            "item": r.get("grain_type", "Wheat"), "wheat": r.get("wheat_weight", 0),
            "output": r.get("output_atta", 0), "method": normalise_method(r.get("payment_method")),
            "deducted": r.get("deducted_flour", 0) or 0,
            "percent": (round((r.get("deducted_flour", 0) or 0) / r["output_atta"] * 100, 2)
                        if r.get("output_atta") else 0),
            "delivered": r.get("final_flour_delivered", r.get("customer_receives", 0)) or 0,
            "grain": f'{r.get("grain_qty", 0)} {r.get("grain_item", "")}'.strip() if r.get("grain_qty") else "",
            "charge": r.get("total_charge", 0), "status": r.get("payment_status", "")} for r in rows]
    return {"columns": [col("date", "Date"), col("invoice", "Invoice"), col("customer", "Customer"),
                        col("item", "Item"), col("wheat", "In (kg)", "qty"), col("output", "Output (kg)", "qty"),
                        col("method", "Paid by"), col("deducted", "Flour kept", "qty"),
                        col("percent", "Deduction %", "qty"), col("delivered", "Delivered", "qty"),
                        col("grain", "Grain taken"), col("charge", "Charge", "money"), col("status", "Status")],
            "rows": out,
            "summary": [{"label": "Grinding charges", "value": round(sum(r["charge"] for r in out), 2), "type": "money"},
                        {"label": "Wheat ground", "value": round(sum(r["wheat"] for r in out), 2)},
                        {"label": "Flour kept as fees", "value": round(sum(r["deducted"] for r in out), 3)},
                        {"label": "Jobs", "value": len(out)}]}

async def rep_party_ledger(start, end, party=None, kind="customer", **_):
    coll = "customers" if kind == "customer" else "suppliers"
    if not party:
        # No party chosen: summarise every one of them instead of an empty page.
        parties = [clean(x) for x in await db[coll].find().sort("name", 1).to_list(5000)]
        rows = []
        for x in parties:
            movements = await party_movements(x["name"])
            deb = sum(m["debit"] for m in movements if in_range(m, start, end))
            cre = sum(m["credit"] for m in movements if in_range(m, start, end))
            bal = round(sum(m["debit"] - m["credit"] for m in movements) + (x.get("opening_balance", 0) or 0), 2)
            if deb or cre or bal:
                rows.append({"party": x["name"], "phone": x.get("phone", ""), "debit": round(deb, 2),
                             "credit": round(cre, 2), "balance": bal})
        return {"columns": [col("party", "Name"), col("phone", "Mobile"), col("debit", "Billed", "money"),
                            col("credit", "Paid", "money"), col("balance", "Balance", "money")],
                "rows": rows,
                "summary": [{"label": "Billed", "value": round(sum(r["debit"] for r in rows), 2), "type": "money"},
                            {"label": "Received" if kind == "customer" else "Paid",
                             "value": round(sum(r["credit"] for r in rows), 2), "type": "money"},
                            {"label": "Outstanding", "value": round(sum(r["balance"] for r in rows), 2), "type": "money"}]}

    master = await db[coll].find_one({"party_key": party_key(party)})
    opening = (master or {}).get("opening_balance", 0) or 0
    movements = await party_movements(party)
    before_start = sum(m["debit"] - m["credit"] for m in movements if str(m.get("date") or "") < start)
    running = round(opening + before_start, 2)
    rows = [{"date": start, "particulars": "Opening balance", "debit": "", "credit": "", "balance": running}]
    for m in movements:
        if not in_range(m, start, end):
            continue
        running = round(running + m["debit"] - m["credit"], 2)
        rows.append({"date": m["date"], "particulars": m["particulars"],
                     "debit": m["debit"] or "", "credit": m["credit"] or "", "balance": running})
    return {"columns": [col("date", "Date"), col("particulars", "Particulars"), col("debit", "Debit", "money"),
                        col("credit", "Credit", "money"), col("balance", "Balance", "money")],
            "rows": rows,
            "summary": [{"label": "Closing balance", "value": running, "type": "money"},
                        {"label": "Entries", "value": max(len(rows) - 1, 0)}]}

async def rep_customer_ledger(start, end, party=None, **kw):
    return await rep_party_ledger(start, end, party, kind="customer")

async def rep_supplier_ledger(start, end, party=None, **kw):
    return await rep_party_ledger(start, end, party, kind="supplier")

async def rep_cash_ledger(start, end, **_):
    settings = await get_settings_doc()
    payments = await db.payments.find().to_list(30000)
    expenses = [clean(e) for e in await db.expenses.find().to_list(20000)]
    cash_rows = [p for p in payments if clean_mode(p.get("payment_mode")) == "Cash"]

    def signed(p):
        # Customer receipts come in; refunds and supplier payments go out.
        amt = p.get("amount", 0) or 0
        return amt if (p.get("party_type") == "customer" and amt >= 0) else -abs(amt)

    opening = round((settings.get("starting_cash", 0) or 0)
                    + sum(signed(p) for p in cash_rows if str(p.get("date") or "") < start)
                    - sum(e.get("amount", 0) for e in expenses if str(e.get("date") or "") < start), 2)
    entries = []
    for p in cash_rows:
        if in_range(p, start, end):
            amt = signed(p)
            entries.append({"date": p.get("date"), "particulars": f'{p.get("party_name", "")} · {p.get("note", "")}'.strip(" ·"),
                            "receipt": round(amt, 2) if amt > 0 else "", "payment": round(-amt, 2) if amt < 0 else ""})
    for e in expenses:
        if in_range(e, start, end):
            entries.append({"date": e.get("date"), "particulars": f'Expense · {e.get("category", "")} {e.get("description", "")}'.strip(),
                            "receipt": "", "payment": round(e.get("amount", 0), 2)})
    entries.sort(key=lambda r: str(r.get("date") or ""))
    running = opening
    rows = [{"date": start, "particulars": "Opening balance", "receipt": "", "payment": "", "balance": running}]
    for e in entries:
        running = round(running + (e["receipt"] or 0) - (e["payment"] or 0), 2)
        rows.append({**e, "balance": running})
    return {"columns": [col("date", "Date"), col("particulars", "Particulars"), col("receipt", "Receipt", "money"),
                        col("payment", "Payment", "money"), col("balance", "Balance", "money")],
            "rows": rows,
            "summary": [{"label": "Opening", "value": opening, "type": "money"},
                        {"label": "Receipts", "value": round(sum(e["receipt"] or 0 for e in entries), 2), "type": "money"},
                        {"label": "Payments", "value": round(sum(e["payment"] or 0 for e in entries), 2), "type": "money"},
                        {"label": "Closing", "value": running, "type": "money"}]}

async def rep_bank_ledger(start, end, bank_id=None, mode=None, **_):
    q = {}
    if bank_id: q["bank_id"] = bank_id
    if mode: q["mode"] = clean_mode(mode)
    txns = [clean(t) for t in await db.bank_txns.find(q).sort("date", 1).to_list(30000)]
    accounts = {a["id"]: a for a in await db.bank_accounts.find().to_list(200)}
    opening_base = sum((a.get("opening_balance", 0) or 0) for a in accounts.values()
                       if not bank_id or a["id"] == bank_id)
    opening = round(opening_base + sum(t.get("amount", 0) for t in txns if str(t.get("date") or "") < start), 2)
    running = opening
    rows = [{"date": start, "bank": "", "particulars": "Opening balance", "mode": "", "reference": "",
             "receipt": "", "payment": "", "balance": running}]
    receipts = payments_out = 0.0
    for t in txns:
        if not in_range(t, start, end):
            continue
        amt = t.get("amount", 0) or 0
        running = round(running + amt, 2)
        if amt >= 0: receipts += amt
        else: payments_out += -amt
        rows.append({"date": t.get("date"), "bank": accounts.get(t.get("bank_id"), {}).get("bank_name", ""),
                     "particulars": f'{t.get("txn_type")} · {t.get("party_name", "")}'.strip(" ·"),
                     "mode": t.get("mode", ""), "reference": t.get("reference", ""),
                     "receipt": round(amt, 2) if amt >= 0 else "", "payment": round(-amt, 2) if amt < 0 else "",
                     "balance": running})
    return {"columns": [col("date", "Date"), col("bank", "Account"), col("particulars", "Particulars"),
                        col("mode", "Mode"), col("reference", "Reference"), col("receipt", "Receipt", "money"),
                        col("payment", "Payment", "money"), col("balance", "Balance", "money")],
            "rows": rows,
            "summary": [{"label": "Opening", "value": opening, "type": "money"},
                        {"label": "Received", "value": round(receipts, 2), "type": "money"},
                        {"label": "Paid", "value": round(payments_out, 2), "type": "money"},
                        {"label": "Closing", "value": running, "type": "money"}]}

async def rep_material_ledger(start, end, party=None, item=None, **_):
    rows = [clean(r) for r in await db.material_ledger.find().sort("date", -1).to_list(30000)]
    rows = [r for r in rows if in_range(r, start, end)]
    if party: rows = [r for r in rows if r.get("party_name") == party]
    if item: rows = [r for r in rows if r.get("item") == item]
    out = [{"date": r.get("date"), "party": r.get("party_name", ""), "item": r.get("item", ""),
            "qty": r.get("qty", 0), "unit": r.get("unit", "kg"), "value": r.get("value", 0),
            "note": r.get("note", "")} for r in rows]
    return {"columns": [col("date", "Date"), col("party", "Party"), col("item", "Item"),
                        col("qty", "Quantity", "qty"), col("unit", "Unit"), col("value", "Value", "money"),
                        col("note", "Particulars")],
            "rows": out,
            "summary": [{"label": "Received in kind", "value": round(sum(r["qty"] for r in out), 3)},
                        {"label": "Value", "value": round(sum(r["value"] for r in out), 2), "type": "money"}]}

async def rep_stock(start, end, item=None, **_):
    rows = await product_movements(start, end)
    if item: rows = [r for r in rows if r["item"] == item]
    return {"columns": [col("item", "Item"), col("category", "Category"), col("unit", "Unit"),
                        col("opening", "Opening", "qty"), col("purchased", "Purchased", "qty"),
                        col("produced", "Produced", "qty"), col("sold", "Sold", "qty"),
                        col("consumed", "Consumed", "qty"), col("adjusted", "Adjusted", "qty"),
                        col("closing", "Closing", "qty"), col("current", "Available now", "qty"),
                        col("value", "Stock value", "money")],
            "rows": rows,
            "summary": [{"label": "Items", "value": len(rows)},
                        {"label": "Stock value", "value": round(sum(r["value"] for r in rows), 2), "type": "money"}]}

async def rep_item_wise(start, end, item=None, **_):
    """Same movements as the stock report, limited to items that actually moved."""
    rows = [r for r in await product_movements(start, end)
            if r["purchased"] or r["sold"] or r["produced"] or r["consumed"] or r["adjusted"]]
    if item: rows = [r for r in rows if r["item"] == item]
    return {"columns": [col("item", "Item"), col("unit", "Unit"), col("opening", "Opening", "qty"),
                        col("purchased", "Purchased", "qty"), col("produced", "Produced", "qty"),
                        col("sold", "Sold", "qty"), col("consumed", "Consumed", "qty"),
                        col("closing", "Closing", "qty"), col("current", "Available now", "qty")],
            "rows": rows,
            "summary": [{"label": "Items moved", "value": len(rows)},
                        {"label": "Sold", "value": round(sum(r["sold"] for r in rows), 3)},
                        {"label": "Purchased", "value": round(sum(r["purchased"] for r in rows), 3)}]}

async def _money_flows(start, end):
    """Receipts and payments in a period, split by how the money moved."""
    payments = await db.payments.find().to_list(30000)
    expenses = [clean(e) for e in await db.expenses.find().to_list(20000)]
    inflow, outflow = {}, {}
    for pay in payments:
        if not in_range(pay, start, end):
            continue
        mode = clean_mode(pay.get("payment_mode"))
        if is_material_mode(mode):
            continue  # settled in kind — no money moved
        amt = pay.get("amount", 0) or 0
        if pay.get("party_type") == "customer" and amt >= 0:
            inflow[mode] = round(inflow.get(mode, 0) + amt, 2)
        else:
            outflow[mode] = round(outflow.get(mode, 0) + abs(amt), 2)
    exp_total = round(sum(e.get("amount", 0) for e in expenses if in_range(e, start, end)), 2)
    outflow["Cash"] = round(outflow.get("Cash", 0) + exp_total, 2)
    return inflow, outflow, exp_total

async def rep_daily_summary(start, end, **_):
    inflow, outflow, exp_total = await _money_flows(start, end)
    cash = await rep_cash_ledger(start, end)
    bank = await rep_bank_ledger(start, end)
    sales = await rep_sales(start, end)
    purch = await rep_purchases(start, end)
    grind = await rep_grinding(start, end)
    rows = [
        {"head": "Sales", "amount": sales["summary"][0]["value"]},
        {"head": "Grinding charges", "amount": grind["summary"][0]["value"]},
        {"head": "Purchases", "amount": purch["summary"][0]["value"]},
        {"head": "Expenses", "amount": exp_total},
        {"head": "Cash received", "amount": inflow.get("Cash", 0)},
        {"head": "Bank received", "amount": round(sum(v for k, v in inflow.items() if k != "Cash"), 2)},
        {"head": "Cash opening", "amount": cash["summary"][0]["value"]},
        {"head": "Cash closing", "amount": cash["summary"][3]["value"]},
        {"head": "Bank opening", "amount": bank["summary"][0]["value"]},
        {"head": "Bank closing", "amount": bank["summary"][3]["value"]},
    ]
    return {"columns": [col("head", "Particulars"), col("amount", "Amount", "money")],
            "rows": rows,
            "summary": [{"label": "Total receipts", "value": round(sum(inflow.values()), 2), "type": "money"},
                        {"label": "Total payments", "value": round(sum(outflow.values()), 2), "type": "money"},
                        {"label": "Cash closing", "value": cash["summary"][3]["value"], "type": "money"},
                        {"label": "Bank closing", "value": bank["summary"][3]["value"], "type": "money"}]}

async def rep_bank_book(start, end, **_):
    """Receipts and payments per bank mode, with the day's opening and closing."""
    inflow, outflow, _ = await _money_flows(start, end)
    bank = await rep_bank_ledger(start, end)
    rows = []
    for mode in BANK_MODES:
        got, paid = inflow.get(mode, 0), outflow.get(mode, 0)
        if got or paid:
            rows.append({"mode": mode, "receipts": got, "payments": paid, "net": round(got - paid, 2)})
    return {"columns": [col("mode", "Mode"), col("receipts", "Receipts", "money"),
                        col("payments", "Payments", "money"), col("net", "Net", "money")],
            "rows": rows,
            "summary": [{"label": "Opening", "value": bank["summary"][0]["value"], "type": "money"},
                        {"label": "Total receipts", "value": round(sum(r["receipts"] for r in rows), 2), "type": "money"},
                        {"label": "Total payments", "value": round(sum(r["payments"] for r in rows), 2), "type": "money"},
                        {"label": "Closing", "value": bank["summary"][3]["value"], "type": "money"}]}

async def rep_outstanding(start, end, **_):
    """Everything still owed, both directions. Ignores the date range on
    purpose: a balance is what it is today, not what it was in a window."""
    rows = []
    for c in await db.customers.find().sort("name", 1).to_list(5000):
        movements = await party_movements(c["name"])
        bal = round(sum(m["debit"] - m["credit"] for m in movements) + (c.get("opening_balance", 0) or 0), 2)
        if abs(bal) >= 0.01:
            rows.append({"party": c["name"], "type": "Customer", "phone": c.get("phone", ""),
                         "receivable": bal if bal > 0 else "", "payable": -bal if bal < 0 else ""})
    for sup in await db.suppliers.find().sort("name", 1).to_list(5000):
        movements = await party_movements(sup["name"])
        bal = round(sum(m["debit"] - m["credit"] for m in movements) + (sup.get("opening_balance", 0) or 0), 2)
        if abs(bal) >= 0.01:
            rows.append({"party": sup["name"], "type": "Supplier", "phone": sup.get("phone", ""),
                         "receivable": -bal if bal < 0 else "", "payable": bal if bal > 0 else ""})
    return {"columns": [col("party", "Party"), col("type", "Type"), col("phone", "Mobile"),
                        col("receivable", "Receivable", "money"), col("payable", "Payable", "money")],
            "rows": rows,
            "summary": [{"label": "Receivable", "value": round(sum(r["receivable"] or 0 for r in rows), 2), "type": "money"},
                        {"label": "Payable", "value": round(sum(r["payable"] or 0 for r in rows), 2), "type": "money"}]}

async def rep_day_book(start, end, **_):
    """Every transaction in the period, in date order, whatever its kind."""
    rows = []
    units = {p["name"]: p.get("unit", "kg") for p in await db.products.find().to_list(2000)}
    for d in await db.sales.find().to_list(20000):
        if in_range(d, start, end):
            rows.append({"date": d.get("date"), "type": "Sale", "reference": d.get("invoice_number", ""),
                         "party": d.get("customer_name", ""),
                         "particulars": f'{d.get("quantity", 0)} {units.get(d.get("product_name"), "")} {d.get("product_name", "")}'.strip(),
                         "amount": d.get("total", 0)})
    for d in await db.purchases.find().to_list(20000):
        if in_range(d, start, end):
            rows.append({"date": d.get("date"), "type": "Purchase", "reference": "",
                         "party": d.get("supplier_name", ""),
                         "particulars": f'{d.get("quantity", 0)} {d.get("product_name", "")}',
                         "amount": d.get("total", 0)})
    for d in await db.grinding.find().to_list(20000):
        if in_range(d, start, end):
            rows.append({"date": d.get("date"), "type": "Grinding", "reference": d.get("invoice_number", ""),
                         "party": d.get("customer_name", ""),
                         "particulars": f'{d.get("wheat_weight", 0)} kg · {normalise_method(d.get("payment_method"))}',
                         "amount": d.get("total_charge", 0)})
    for d in await db.oil.find().to_list(20000):
        if in_range(d, start, end):
            rows.append({"date": d.get("date"), "type": "Oil", "reference": d.get("invoice_number", ""),
                         "party": d.get("customer_name", ""),
                         "particulars": f'{d.get("quantity_received", 0)} kg {d.get("seed_type", "")}',
                         "amount": d.get("total", 0)})
    for d in await db.exchanges.find().to_list(20000):
        if in_range(d, start, end):
            rows.append({"date": d.get("date"), "type": "Exchange", "reference": d.get("invoice_number", ""),
                         "party": d.get("customer_name", ""),
                         "particulars": f'{d.get("wheat_qty", 0)} kg wheat for flour',
                         "amount": d.get("grinding_charge", 0)})
    for d in await db.expenses.find().to_list(20000):
        if in_range(d, start, end):
            rows.append({"date": d.get("date"), "type": "Expense", "reference": "", "party": "",
                         "particulars": f'{d.get("category", "")} {d.get("description", "")}'.strip(),
                         "amount": d.get("amount", 0)})
    rows.sort(key=lambda r: (str(r.get("date") or ""), r["type"]))
    income = sum(r["amount"] for r in rows if r["type"] in ("Sale", "Grinding", "Oil", "Exchange"))
    return {"columns": [col("date", "Date"), col("type", "Type"), col("reference", "Reference"),
                        col("party", "Party"), col("particulars", "Particulars"), col("amount", "Amount", "money")],
            "rows": rows,
            "summary": [{"label": "Entries", "value": len(rows)},
                        {"label": "Income", "value": round(income, 2), "type": "money"},
                        {"label": "Purchases", "value": round(sum(r["amount"] for r in rows if r["type"] == "Purchase"), 2), "type": "money"},
                        {"label": "Expenses", "value": round(sum(r["amount"] for r in rows if r["type"] == "Expense"), 2), "type": "money"}]}

async def rep_profit_loss(start, end, **_):
    sales = [d for d in await db.sales.find().to_list(30000) if in_range(d, start, end)]
    grind = [d for d in await db.grinding.find().to_list(30000) if in_range(d, start, end)]
    oil = [d for d in await db.oil.find().to_list(30000) if in_range(d, start, end)]
    exch = [d for d in await db.exchanges.find().to_list(30000) if in_range(d, start, end)]
    expenses = [d for d in await db.expenses.find().to_list(30000) if in_range(d, start, end)]
    products = {p["name"]: p for p in await db.products.find().to_list(2000)}
    cost_by_name = {n: p.get("cost_per_unit", 0) or 0 for n, p in products.items()}

    sale_income = round(sum(d.get("total", 0) for d in sales), 2)
    cogs = round(sum(sale_cogs(d, cost_by_name) for d in sales), 2)
    grind_income = round(sum(d.get("total_charge", 0) for d in grind), 2)
    oil_income = round(sum(d.get("total", 0) for d in oil), 2)
    exch_income = round(sum(d.get("grinding_charge", 0) for d in exch), 2)
    exp_total = round(sum(d.get("amount", 0) for d in expenses), 2)

    by_cat = {}
    for e in expenses:
        by_cat[e.get("category", "Other")] = round(by_cat.get(e.get("category", "Other"), 0) + e.get("amount", 0), 2)

    rows = [{"head": "Sales income", "amount": sale_income},
            {"head": "Less: cost of goods sold", "amount": -cogs},
            {"head": "Gross profit on sales", "amount": round(sale_income - cogs, 2)},
            {"head": "Grinding charges", "amount": grind_income},
            {"head": "Oil extraction charges", "amount": oil_income},
            {"head": "Exchange grinding charges", "amount": exch_income}]
    rows += [{"head": f'Expense · {k}', "amount": -v} for k, v in sorted(by_cat.items())]
    net = round(sale_income - cogs + grind_income + oil_income + exch_income - exp_total, 2)
    rows.append({"head": "Net profit" if net >= 0 else "Net loss", "amount": net})
    return {"columns": [col("head", "Particulars"), col("amount", "Amount", "money")],
            "rows": rows,
            "summary": [{"label": "Income", "value": round(sale_income + grind_income + oil_income + exch_income, 2), "type": "money"},
                        {"label": "Cost of goods sold", "value": cogs, "type": "money"},
                        {"label": "Expenses", "value": exp_total, "type": "money"},
                        {"label": "Net profit" if net >= 0 else "Net loss", "value": net, "type": "money"}]}

async def rep_trial_balance(start, end, **_):
    """A trial balance over the accounts this app actually keeps.

    The app has no chart of accounts, so this is built from the real balances it
    does hold — cash, each bank account, party balances, stock at cost, and the
    income and expense totals for the period — rather than inventing ledgers it
    does not maintain.
    """
    cash = await rep_cash_ledger(start, end)
    rows = [{"account": "Cash in hand", "debit": max(cash["summary"][3]["value"], 0),
             "credit": max(-cash["summary"][3]["value"], 0)}]
    for a in await db.bank_accounts.find().sort("bank_name", 1).to_list(200):
        bal = await bank_balance(a["id"])
        rows.append({"account": f'Bank · {a.get("bank_name")}', "debit": max(bal, 0), "credit": max(-bal, 0)})

    out = await rep_outstanding(start, end)
    receivable = out["summary"][0]["value"]
    payable = out["summary"][1]["value"]
    rows.append({"account": "Accounts receivable", "debit": receivable, "credit": 0})
    rows.append({"account": "Accounts payable", "debit": 0, "credit": payable})

    stock = await rep_stock(start, end)
    rows.append({"account": "Closing stock (at cost)", "debit": stock["summary"][1]["value"], "credit": 0})

    pl = await rep_profit_loss(start, end)
    rows.append({"account": "Income", "debit": 0, "credit": pl["summary"][0]["value"]})
    rows.append({"account": "Cost of goods sold", "debit": pl["summary"][1]["value"], "credit": 0})
    rows.append({"account": "Expenses", "debit": pl["summary"][2]["value"], "credit": 0})

    td = round(sum(r["debit"] for r in rows), 2)
    tc = round(sum(r["credit"] for r in rows), 2)
    return {"columns": [col("account", "Account"), col("debit", "Debit", "money"), col("credit", "Credit", "money")],
            "rows": rows,
            "summary": [{"label": "Total debit", "value": td, "type": "money"},
                        {"label": "Total credit", "value": tc, "type": "money"},
                        {"label": "Difference", "value": round(td - tc, 2), "type": "money"}],
            "note": ("Built from the balances this app keeps — cash, banks, party balances, stock at cost "
                     "and the period's income and expenses. It is not a double-entry trial balance, so the "
                     "two sides are not expected to agree exactly.")}

# Registry. `party` says which picker the filter bar should offer, so the
# frontend does not need its own table of which report takes what.
REPORTS = {
    "daily-summary":    {"title": "Daily Transaction Summary", "group": "Daily",   "fn": "rep_daily_summary"},
    "cash-book":        {"title": "Cash Book",                 "group": "Daily",   "fn": "rep_cash_ledger"},
    "bank-book":        {"title": "Bank Book",                 "group": "Daily",   "fn": "rep_bank_book"},
    "day-book":         {"title": "Day Book",                  "group": "Daily",   "fn": "rep_day_book"},
    "item-wise":        {"title": "Item-wise Movement",        "group": "Stock",   "fn": "rep_item_wise", "item": True},
    "stock":            {"title": "Stock Report",              "group": "Stock",   "fn": "rep_stock",     "item": True},
    "sales":            {"title": "Sales Report",              "group": "Trading", "fn": "rep_sales",     "party": "customer", "item": True},
    "purchases":        {"title": "Purchase Report",           "group": "Trading", "fn": "rep_purchases", "party": "supplier", "item": True},
    "grinding":         {"title": "Grinding Report",           "group": "Trading", "fn": "rep_grinding",  "party": "customer"},
    "customer-ledger":  {"title": "Customer Ledger",           "group": "Ledgers", "fn": "rep_customer_ledger", "party": "customer"},
    "supplier-ledger":  {"title": "Supplier Ledger",           "group": "Ledgers", "fn": "rep_supplier_ledger", "party": "supplier"},
    "cash-ledger":      {"title": "Cash Ledger",               "group": "Ledgers", "fn": "rep_cash_ledger"},
    "bank-ledger":      {"title": "Bank Ledger",               "group": "Ledgers", "fn": "rep_bank_ledger", "bank": True},
    "material-ledger":  {"title": "Grain / Material Ledger",   "group": "Ledgers", "fn": "rep_material_ledger", "party": "customer", "item": True},
    "outstanding":      {"title": "Outstanding (Dues)",        "group": "Summary", "fn": "rep_outstanding"},
    "profit-loss":      {"title": "Profit & Loss",             "group": "Summary", "fn": "rep_profit_loss"},
    "trial-balance":    {"title": "Trial Balance",             "group": "Summary", "fn": "rep_trial_balance"},
}

@api_router.get("/reports")
async def list_reports(user: dict = Depends(get_current_user)):
    return [{"key": k, "title": v["title"], "group": v["group"],
             "party": v.get("party"), "item": bool(v.get("item")), "bank": bool(v.get("bank"))}
            for k, v in REPORTS.items()]

async def run_report(key: str, preset=None, start=None, end=None, party=None,
                     item=None, bank_id=None, mode=None):
    spec = REPORTS.get(key)
    if not spec:
        raise HTTPException(status_code=404, detail=f"Unknown report: {key}")
    s_date, e_date = date_range(preset, start, end)
    data = await globals()[spec["fn"]](s_date, e_date, party=party, item=item,
                                       bank_id=bank_id, mode=mode)
    return {"key": key, "title": spec["title"], "start": s_date, "end": e_date,
            "filters": {"party": party, "item": item, "bank_id": bank_id, "mode": mode},
            **data}

@api_router.get("/reports/{key}")
async def get_report(key: str, preset: Optional[str] = None, start: Optional[str] = None,
                     end: Optional[str] = None, party: Optional[str] = None,
                     item: Optional[str] = None, bank_id: Optional[str] = None,
                     mode: Optional[str] = None, user: dict = Depends(get_current_user)):
    return await run_report(key, preset, start, end, party, item, bank_id, mode)

def _fmt_cell(value, kind):
    if value is None or value == "":
        return ""
    if kind in ("money", "qty") and isinstance(value, (int, float)):
        return f'{value:,.2f}' if kind == "money" else f'{value:g}'
    return str(value)

def report_to_xlsx(report: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    # Excel rejects : \ / ? * [ ] in a sheet name and caps it at 31 characters,
    # which "Grain / Material Ledger" falls foul of.
    safe = report["title"]
    for ch in ':\\/?*[]':
        safe = safe.replace(ch, "-")
    ws.title = safe[:31].strip() or "Report"
    ws.append([report["title"]])
    ws.append([f'{report["start"]} to {report["end"]}'])
    applied = [f'{k}: {v}' for k, v in (report.get("filters") or {}).items() if v]
    if applied:
        ws.append([" · ".join(applied)])
    ws.append([])
    ws.append([c["label"] for c in report["columns"]])
    for row in report["rows"]:
        # Numbers stay numbers so Excel can sum them; only blanks are coerced.
        ws.append([row.get(c["key"], "") if row.get(c["key"], "") != "" else "" for c in report["columns"]])
    ws.append([])
    for item in report.get("summary", []):
        ws.append([item["label"], item["value"]])
    for i, c in enumerate(report["columns"], start=1):
        width = max(len(c["label"]) + 2,
                    *(len(_fmt_cell(r.get(c["key"]), c["type"])) + 2 for r in report["rows"])) if report["rows"] else len(c["label"]) + 2
        ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = min(width, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def report_to_pdf(report: dict) -> io.BytesIO:
    buf = io.BytesIO()
    # Landscape: these tables are wider than a portrait page can hold.
    doc = SimpleDocTemplate(buf, pagesize=(A4[1], A4[0]), leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Heading1"], fontSize=15, spaceAfter=2)
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    story = [Paragraph("Gangotri Flour &amp; Oil Mill", title),
             Paragraph(f'<b>{report["title"]}</b> &nbsp; {report["start"]} to {report["end"]}', styles["Normal"])]
    applied = [f'{k}: {v}' for k, v in (report.get("filters") or {}).items() if v]
    if applied:
        story.append(Paragraph(" · ".join(applied), small))
    story.append(Spacer(1, 6))

    head = [c["label"] for c in report["columns"]]
    body = [[_fmt_cell(r.get(c["key"]), c["type"]) for c in report["columns"]] for r in report["rows"]]
    if not body:
        body = [["No entries for this period"] + [""] * (len(head) - 1)]
    tbl = Table([head] + body, repeatRows=1)
    style = [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f5132")),
             ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
             ("FONTSIZE", (0, 0), (-1, -1), 7.5),
             ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
             ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f6f6")])]
    for i, c in enumerate(report["columns"]):
        if c["align"] == "right":
            style.append(("ALIGN", (i, 0), (i, -1), "RIGHT"))
    tbl.setStyle(TableStyle(style))
    story.append(tbl)

    if report.get("summary"):
        story.append(Spacer(1, 8))
        srows = [[s["label"], _fmt_cell(s["value"], s.get("type", "text"))] for s in report["summary"]]
        st = Table(srows, colWidths=[60*mm, 40*mm])
        st.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9),
                                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        story.append(st)
    if report.get("note"):
        story.append(Spacer(1, 6))
        story.append(Paragraph(report["note"], small))
    doc.build(story)
    buf.seek(0)
    return buf

@api_router.get("/reports/{key}/export")
async def export_report(key: str, request: Request, format: str = "xlsx",
                        preset: Optional[str] = None, start: Optional[str] = None,
                        end: Optional[str] = None, party: Optional[str] = None,
                        item: Optional[str] = None, bank_id: Optional[str] = None,
                        mode: Optional[str] = None):
    # Downloads open in a new tab, which cannot set an Authorization header, so
    # the cookie is checked directly here rather than through the dependency.
    await get_current_user(request)
    report = await run_report(key, preset, start, end, party, item, bank_id, mode)
    stamp = f'{report["start"]}_to_{report["end"]}'
    if format == "pdf":
        return StreamingResponse(report_to_pdf(report), media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{key}_{stamp}.pdf"'})
    return StreamingResponse(report_to_xlsx(report),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{key}_{stamp}.xlsx"'})

# ==================== Global search ====================
# Reuses the report shape — columns, rows, summary — so results print and
# export through the same PDF and Excel writers rather than a second set.
#
# Each scope declares which collection it reads, which fields the free-text box
# matches, and how to turn a row into something the UI can open, print or
# delete. Adding a searchable module is one entry.

SEARCH_SCOPES = {
    "sales": {
        "title": "Sales", "coll": "sales", "route": "/sales", "invoice": True,
        "text": ("invoice_number", "customer_name", "product_name"),
        "party": "customer_name", "item": "product_name",
        "columns": [col("date", "Date"), col("invoice_number", "Invoice"), col("customer_name", "Customer"),
                    col("product_name", "Item"), col("quantity", "Qty", "qty"),
                    col("total", "Total", "money"), col("balance_due", "Balance", "money"),
                    col("payment_status", "Status")],
    },
    "purchases": {
        "title": "Purchases", "coll": "purchases", "route": "/inventory",
        "text": ("supplier_name", "product_name"),
        "party": "supplier_name", "item": "product_name",
        "columns": [col("date", "Date"), col("supplier_name", "Supplier"), col("product_name", "Item"),
                    col("quantity", "Qty", "qty"), col("total", "Total", "money"),
                    col("balance_due", "Balance", "money"), col("payment_status", "Status")],
    },
    "grinding": {
        "title": "Grinding", "coll": "grinding", "route": "/grinding", "invoice": True,
        "text": ("invoice_number", "customer_name", "grain_type"),
        "party": "customer_name",
        "columns": [col("date", "Date"), col("invoice_number", "Invoice"), col("customer_name", "Customer"),
                    col("grain_type", "Item"), col("wheat_weight", "In (kg)", "qty"),
                    col("payment_method", "Paid by"), col("total_charge", "Charge", "money"),
                    col("payment_status", "Status")],
    },
    "oil": {
        "title": "Oil extraction", "coll": "oil", "route": "/oil", "invoice": True,
        "text": ("invoice_number", "customer_name", "seed_type"),
        "party": "customer_name",
        "columns": [col("date", "Date"), col("invoice_number", "Invoice"), col("customer_name", "Customer"),
                    col("seed_type", "Seed"), col("quantity_received", "Received", "qty"),
                    col("total", "Total", "money"), col("payment_status", "Status")],
    },
    "exchanges": {
        "title": "Exchange", "coll": "exchanges", "route": "/exchange", "invoice": True,
        "text": ("invoice_number", "customer_name"),
        "party": "customer_name",
        "columns": [col("date", "Date"), col("invoice_number", "Invoice"), col("customer_name", "Customer"),
                    col("wheat_qty", "Wheat", "qty"), col("final_flour_delivered", "Delivered", "qty"),
                    col("grinding_charge", "Charge", "money"), col("payment_method", "Paid by")],
    },
    "payments": {
        "title": "Payments", "coll": "payments", "route": "/customers",
        "text": ("party_name", "note", "reference"),
        "party": "party_name",
        "columns": [col("date", "Date"), col("party_name", "Party"), col("party_type", "Type"),
                    col("payment_mode", "Mode"), col("amount", "Amount", "money"), col("note", "Note")],
    },
    "bank": {
        "title": "Bank transactions", "coll": "bank_txns", "route": "/banks",
        "text": ("party_name", "reference", "note", "txn_type"),
        "columns": [col("date", "Date"), col("txn_type", "Type"), col("mode", "Mode"),
                    col("party_name", "Party"), col("reference", "Reference"),
                    col("amount", "Amount", "money")],
    },
    "expenses": {
        "title": "Expenses", "coll": "expenses", "route": "/expenses",
        "text": ("category", "description"),
        "columns": [col("date", "Date"), col("category", "Category"),
                    col("description", "Description"), col("amount", "Amount", "money")],
    },
    "production": {
        "title": "Production", "coll": "production", "route": "/production",
        "text": ("mill", "input_product_name"),
        "item": "input_product_name",
        "columns": [col("date", "Date"), col("mill", "Mill"), col("input_product_name", "Input"),
                    col("input_quantity", "Qty", "qty"), col("input_cost", "Cost", "money")],
    },
    "customers": {
        "title": "Customers", "coll": "customers", "route": "/customers", "undated": True,
        "text": ("name", "phone", "gstin", "pan_aadhaar", "address"),
        "columns": [col("name", "Name"), col("phone", "Mobile"), col("address", "Address"),
                    col("gstin", "GSTIN"), col("opening_balance", "Opening", "money")],
    },
    "suppliers": {
        "title": "Suppliers", "coll": "suppliers", "route": "/suppliers", "undated": True,
        "text": ("name", "phone", "gstin", "pan_aadhaar", "address"),
        "columns": [col("name", "Name"), col("phone", "Mobile"), col("address", "Address"),
                    col("gstin", "GSTIN"), col("opening_balance", "Opening", "money")],
    },
    "inventory": {
        "title": "Stock", "coll": "products", "route": "/inventory", "undated": True,
        "text": ("name", "category"), "item": "name",
        "columns": [col("name", "Item"), col("category", "Category"), col("unit", "Unit"),
                    col("current_stock", "In stock", "qty"), col("rate", "Rate", "money")],
    },
}

# Which field carries the money status, per scope, so "Due" means the same
# thing everywhere. Bank rows and masters have none.
def matches_status(scope: str, row: dict, status: str) -> bool:
    if not status:
        return True
    current = row.get("payment_status")
    if current is None:
        return False
    if status.lower() in ("due", "pending", "unpaid"):
        return current in ("Pending", "Partial")
    return current.lower() == status.lower()

def matches_mode(row: dict, mode: str) -> bool:
    if not mode:
        return True
    asked = normalise_method(mode)
    # Settled in kind, so match only the method. clean_mode returns Cash for
    # anything it does not recognise, which would otherwise make a search for
    # "Flour Deduction" return every cash job as well.
    if asked in KIND_METHODS:
        return bool(row.get("payment_method")) and normalise_method(row["payment_method"]) == asked
    want = clean_mode(mode)
    for field in ("payment_mode", "mode"):
        if row.get(field) and clean_mode(row[field]) == want:
            return True
    # A money method may be recorded on the bill rather than on a payment row.
    pm = row.get("payment_method")
    if pm and normalise_method(pm) not in KIND_METHODS and clean_mode(pm) == want:
        return True
    return False

async def run_search(q="", scopes=None, preset=None, start=None, end=None, party=None,
                     item=None, mode=None, status=None, limit=200):
    s_date, e_date = date_range(preset, start, end)
    needle = (q or "").strip().lower()
    wanted = [k for k in (scopes or SEARCH_SCOPES.keys()) if k in SEARCH_SCOPES]
    dated = bool(preset or start or end)
    groups = []

    for key in wanted:
        spec = SEARCH_SCOPES[key]
        query = {}
        # Push the date filter into Mongo, where it is indexed, so the text
        # match below only ever runs over the rows in the period.
        if dated and not spec.get("undated"):
            query["date"] = {"$gte": s_date, "$lte": e_date}
        if party and spec.get("party"):
            query[spec["party"]] = party
        if item and spec.get("item"):
            query[spec["item"]] = item

        rows = [clean(r) for r in await db[spec["coll"]].find(query).to_list(20000)]

        if needle:
            rows = [r for r in rows
                    if any(needle in str(r.get(f, "") or "").lower() for f in spec["text"])]
        if mode:
            rows = [r for r in rows if matches_mode(r, mode)]
        if status:
            rows = [r for r in rows if matches_status(key, r, status)]

        rows.sort(key=lambda r: str(r.get("date") or r.get("name") or ""), reverse=not spec.get("undated"))
        total = len(rows)
        rows = rows[:limit]
        for r in rows:
            # Everything the results table needs to open, print or delete a row
            # without knowing which module it came from.
            r["_scope"] = key
            r["_route"] = spec["route"]
            r["_invoice"] = bool(spec.get("invoice")) and bool(r.get("invoice_number"))
            if key == "grinding" and r.get("payment_method"):
                r["payment_method"] = normalise_method(r["payment_method"])
        if total:
            groups.append({"scope": key, "title": spec["title"], "count": total,
                           "shown": len(rows), "columns": spec["columns"], "rows": rows})

    return {"query": q, "start": s_date if dated else None, "end": e_date if dated else None,
            "dated": dated, "total": sum(g["count"] for g in groups), "groups": groups}

@api_router.get("/search")
async def global_search(q: str = "", scope: Optional[str] = None, preset: Optional[str] = None,
                        start: Optional[str] = None, end: Optional[str] = None,
                        party: Optional[str] = None, item: Optional[str] = None,
                        mode: Optional[str] = None, status: Optional[str] = None,
                        limit: int = 200, user: dict = Depends(get_current_user)):
    scopes = [s.strip() for s in scope.split(",")] if scope else None
    return await run_search(q, scopes, preset, start, end, party, item, mode, status, limit)

@api_router.get("/search/scopes")
async def search_scopes(user: dict = Depends(get_current_user)):
    return [{"key": k, "title": v["title"], "undated": bool(v.get("undated"))}
            for k, v in SEARCH_SCOPES.items()]

@api_router.get("/search/export")
async def export_search(request: Request, format: str = "xlsx", q: str = "",
                        scope: Optional[str] = None, preset: Optional[str] = None,
                        start: Optional[str] = None, end: Optional[str] = None,
                        party: Optional[str] = None, item: Optional[str] = None,
                        mode: Optional[str] = None, status: Optional[str] = None):
    await get_current_user(request)
    scopes = [s.strip() for s in scope.split(",")] if scope else None
    found = await run_search(q, scopes, preset, start, end, party, item, mode, status, limit=20000)
    # Flattened into one table with a Section column, so a mixed result set
    # exports as a single readable sheet rather than one file per scope.
    columns = [col("_section", "Section"), col("date", "Date"), col("_summary", "Details"),
               col("_amount", "Amount", "money")]
    rows = []
    for g in found["groups"]:
        for r in g["rows"]:
            parts = [str(r.get(c["key"], "")) for c in g["columns"]
                     if c["type"] == "text" and r.get(c["key"]) not in (None, "")]
            amount = next((r.get(c["key"]) for c in g["columns"] if c["type"] == "money"), "")
            rows.append({"_section": g["title"], "date": r.get("date", ""),
                         "_summary": " · ".join(parts[:4]), "_amount": amount})
    report = {"key": "search", "title": f'Search results{f" for {q}" if q else ""}',
              "start": found["start"] or "all", "end": found["end"] or "dates",
              "filters": {k: v for k, v in (("party", party), ("item", item),
                                            ("mode", mode), ("status", status)) if v},
              "columns": columns, "rows": rows,
              "summary": [{"label": g["title"], "value": g["count"]} for g in found["groups"]]
                         + [{"label": "Total matches", "value": found["total"]}]}
    if format == "pdf":
        return StreamingResponse(report_to_pdf(report), media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="search_results.pdf"'})
    return StreamingResponse(report_to_xlsx(report),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="search_results.xlsx"'})

# ---------------- Cost basis repair ----------------
# PUT /products used to write cost_per_unit from the request body, and the edit
# form never sent that field, so Pydantic's default of 0 was saved on every
# product edit. A zero cost makes a sale report its full price as profit. The
# fix stopped new damage; this rebuilds what was already lost.

async def cost_inflows(name: str, pid: str):
    """Everything the shop acquired of one product, and what it paid.

    Only inflows carry cost. Sales and production consumption take stock out at
    whatever the average already was, so they cannot help rebuild it.
    """
    lots = []
    for d in await db.purchases.find({"$or": [{"product_id": pid}, {"product_name": name}]}).to_list(20000):
        qty = d.get("quantity", 0) or 0
        if qty > 0:
            lots.append((qty, d.get("total", 0) or 0, "purchase"))
    for d in await db.production.find().to_list(20000):
        for o in d.get("outputs", []):
            if (o.get("product_id") == pid or o.get("product_name") == name) and (o.get("quantity", 0) or 0) > 0:
                lots.append((o["quantity"], o.get("cost", 0) or 0, "production"))
    for d in await db.oil.find().to_list(20000):
        # Cake bought from the customer is stock the shop paid for.
        if d.get("cake_sold_to_shop") and name == "Mustard Oil Cake":
            lots.append((d["cake_sold_to_shop"], (d["cake_sold_to_shop"]) * (d.get("cake_rate", 0) or 0), "cake purchase"))
    for d in await db.grinding.find().to_list(20000):
        method = normalise_method(d.get("payment_method"))
        if method == GRAIN_DEDUCTION and (d.get("grain_item") or "") == name and d.get("grain_qty"):
            lots.append((d["grain_qty"], d.get("grain_value", 0) or 0, "grain fee"))
        elif method == FLOUR_DEDUCTION and grinding_output_name(d) == name and d.get("deducted_flour"):
            lots.append((d["deducted_flour"], d.get("flour_value", 0) or 0, "flour fee"))
    return lots

async def rebuild_costs(only_zero: bool = True, apply: bool = False):
    products = [clean(p) for p in await db.products.find().sort("name", 1).to_list(2000)]
    changes = []
    for prod in products:
        current = prod.get("cost_per_unit", 0) or 0
        if only_zero and current > 0:
            continue
        lots = await cost_inflows(prod["name"], prod["id"])
        qty = sum(q for q, _, _ in lots)
        value = sum(v for _, v, _ in lots)
        if qty <= 0:
            continue
        rebuilt = round(value / qty, 4)
        if abs(rebuilt - current) < 0.0001:
            continue
        changes.append({"id": prod["id"], "name": prod["name"], "unit": prod.get("unit", "kg"),
                        "current_cost": round(current, 4), "rebuilt_cost": rebuilt,
                        "acquired_qty": round(qty, 3), "acquired_value": round(value, 2),
                        "sources": sorted({src for _, _, src in lots}),
                        "stock": prod.get("current_stock", 0),
                        "stock_value_change": round((rebuilt - current) * (prod.get("current_stock", 0) or 0), 2)})
        # Sales snapshot their cost when they are made, so every sale during the
        # corrupted window carries cogs: 0 and still reports its full price as
        # profit. Repairing the product alone would leave the reports wrong.
        # Only zero snapshots are re-stamped: a real one must not be rewritten
        # by a later purchase, which is why the snapshot exists at all.
        broken = [clean(x) for x in await db.sales.find({
            "$or": [{"product_id": prod["id"]}, {"product_name": prod["name"]}],
            "$and": [{"$or": [{"cogs": 0}, {"cogs": None}, {"cogs": {"$exists": False}}]}],
        }).to_list(20000)]
        broken = [b for b in broken if (b.get("quantity", 0) or 0) > 0]
        changes[-1]["sales_restamped"] = len(broken)
        changes[-1]["profit_correction"] = round(sum((b.get("quantity", 0) or 0) * rebuilt for b in broken), 2)

        if apply:
            await db.products.update_one({"id": prod["id"]}, {"$set": {"cost_per_unit": rebuilt}})
            for b in broken:
                qty = b.get("quantity", 0) or 0
                await db.sales.update_one({"id": b["id"]}, {"$set": {
                    "unit_cost": rebuilt, "cogs": round(qty * rebuilt, 2)}})
    return changes

@api_router.get("/products/cost-repair/preview")
async def preview_cost_repair(only_zero: bool = True, user: dict = Depends(require_admin)):
    changes = await rebuild_costs(only_zero=only_zero, apply=False)
    return {"changes": changes, "count": len(changes),
            "stock_value_change": round(sum(c["stock_value_change"] for c in changes), 2),
            "sales_restamped": sum(c.get("sales_restamped", 0) for c in changes),
            "profit_correction": round(sum(c.get("profit_correction", 0) for c in changes), 2),
            "only_zero": only_zero}

@api_router.post("/products/cost-repair")
async def apply_cost_repair(only_zero: bool = True, user: dict = Depends(require_admin)):
    changes = await rebuild_costs(only_zero=only_zero, apply=True)
    for c in changes:
        await log_audit(user, "Rebuilt cost basis",
                        f'{c["name"]}: Rs {c["current_cost"]} → {c["rebuilt_cost"]}/{c["unit"]} '
                        f'from {c["acquired_qty"]} acquired')
    return {"changes": changes, "count": len(changes),
            "stock_value_change": round(sum(c["stock_value_change"] for c in changes), 2),
            "sales_restamped": sum(c.get("sales_restamped", 0) for c in changes),
            "profit_correction": round(sum(c.get("profit_correction", 0) for c in changes), 2)}

@api_router.get("/sales-analytics")
async def sales_analytics(user: dict = Depends(get_current_user)):
    sales = await db.sales.find().to_list(30000)
    products = {p["name"]: p for p in await db.products.find().to_list(1000)}
    t = datetime.now().strftime("%Y-%m-%d")
    m = t[:7]
    y = t[:4]
    def rev(prefix):
        return round(sum(s.get("total", 0) for s in sales if str(s.get("date", "")).startswith(prefix)), 2)
    # Same costing rule as the dashboard, so the two profit figures agree.
    cost_by_name = {n: p.get("cost_per_unit", 0) or 0 for n, p in products.items()}
    agg = {}
    for s in sales:
        n = s.get("product_name", "?")
        a = agg.setdefault(n, {"name": n, "qty": 0.0, "revenue": 0.0, "cogs": 0.0, "profit": 0.0})
        a["qty"] += s.get("quantity", 0)
        a["revenue"] += s.get("total", 0)
        cogs = sale_cogs(s, cost_by_name)
        a["cogs"] += cogs
        a["profit"] += s.get("total", 0) - cogs
    rows = sorted(agg.values(), key=lambda r: r["revenue"], reverse=True)
    for r in rows:
        r["qty"] = round(r["qty"], 2)
        r["revenue"] = round(r["revenue"], 2)
        r["cogs"] = round(r["cogs"], 2)
        r["profit"] = round(r["profit"], 2)
    return {"today": rev(t), "month": rev(m), "year": rev(y),
            "total_revenue": round(sum(s.get("total", 0) for s in sales), 2),
            "total_cogs": round(sum(r["cogs"] for r in rows), 2),
            "total_profit": round(sum(r["profit"] for r in rows), 2),
            "by_product": rows, "top": rows[:5], "least": rows[-5:][::-1] if rows else []}

@api_router.get("/health")
async def health():
    return {"status": "ok"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000"), "http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------- Static frontend ----------------
# In production the React build is copied to backend/static. Serving it from this
# same process keeps the app single-origin, which matters because auth is
# cookie-only: on a split frontend/backend deploy the auth cookie would be a
# third-party cookie and Safari would drop it. No-op in local dev, where CRA
# serves the frontend on :3000 and the block below is skipped.
STATIC_DIR = ROOT_DIR / "static"

if STATIC_DIR.is_dir():
    app.mount("/static", StaticFiles(directory=STATIC_DIR / "static"), name="assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        # Unmatched /api paths are a 404, not the SPA shell.
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not found")
        index = STATIC_DIR / "index.html"
        if full_path:
            candidate = (STATIC_DIR / full_path).resolve()
            # Keep path traversal (../../etc/passwd) inside the build directory.
            if candidate.is_file() and candidate.is_relative_to(STATIC_DIR.resolve()):
                return FileResponse(candidate)
        return FileResponse(index)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
